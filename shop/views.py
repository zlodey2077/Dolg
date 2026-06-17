import io
import json
import logging
from pathlib import Path

from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

PRODUCTS_PER_PAGE = 24  # 4×6 grid; страница меньше 1 МБ HTML на текущих 72 товарах

from .component_validation import (
    missing_spice_model_warning,
    nominal_mismatch_warning,
    unique_warnings,
)
from .models import CartItem, Category, Product
from .services.catalog_filters import (
    CATALOG_FILTERS,
    apply_catalog_filters,
    build_active_filter_tags,
    build_filter_options,
    clean_params,
    compute_catalog_facets,
    hidden_filter_inputs,
    querystring_with,
)
from .smart_search import smart_search


def _log_bom_project_event(request, payload, event_type, event_payload):
    project_id = (
        (payload.get('project') or {}).get('id') if isinstance(payload.get('project'), dict) else None
    ) or payload.get('project_id')
    if not project_id:
        return
    try:
        from Dolg_APP.models import ProjectEvent, SchematicProject

        project = SchematicProject.all_objects.get(pk=project_id)
        event = ProjectEvent.log(
            project=project,
            user=request.user if getattr(request.user, 'is_authenticated', False) else None,
            event_type=event_type,
            payload=event_payload,
        )
        if event:
            try:
                from asgiref.sync import async_to_sync
                from channels.layers import get_channel_layer

                layer = get_channel_layer()
                if layer is not None:
                    async_to_sync(layer.group_send)(
                        f'project-{project.id}',
                        {
                            'type': 'project.event',
                            'event': {
                                'id': event.id,
                                'project_id': project.id,
                                'event_type': event.event_type,
                                'event_label': event.get_event_type_display(),
                                'payload': event.payload,
                                'user': event.user.username if event.user_id else '',
                                'created': event.created_at.isoformat(),
                            },
                        },
                    )
            except Exception:
                pass
    except Exception:
        pass


def get_session_id(request):
    if not request.session.session_key:
        request.session.create()
    return request.session.session_key


def _cart_owner_filter(request):
    """Возвращает Q-подобный dict для фильтрации CartItem.
    - Auth → {'user': request.user}
    - Guest → {'session_id': session_key, 'user__isnull': True}
    """
    if request.user.is_authenticated:
        return {'user': request.user}
    return {'session_id': get_session_id(request), 'user__isnull': True}


def _cart_owner_create_defaults(request):
    """Поля для создания CartItem с правильной привязкой."""
    if request.user.is_authenticated:
        return {'user': request.user, 'session_id': ''}
    return {'user': None, 'session_id': get_session_id(request)}


REB_SLUGS = {'resistors', 'capacitors', 'transistors', 'ics', 'diodes', 'inductors', 'connectors', 'relays'}

MANUFACTURER_RESOURCE_URLS = {
    'yageo': 'https://www.yageo.com/',
    'vishay': 'https://www.vishay.com/',
    'murata': 'https://www.murata.com/',
    'tdk': 'https://product.tdk.com/',
    'kemet': 'https://www.kemet.com/',
    'st': 'https://www.st.com/',
    'nxp': 'https://www.nxp.com/',
    'ti': 'https://www.ti.com/',
    'infineon': 'https://www.infineon.com/',
    'onsemi': 'https://www.onsemi.com/',
    'bourns': 'https://www.bourns.com/',
    'panasonic': 'https://industrial.panasonic.com/',
    'nichicon': 'https://www.nichicon.co.jp/english/',
    'wurth': 'https://www.we-online.com/',
    'apple': 'https://www.apple.com/',
    'samsung': 'https://www.samsung.com/',
    'sony': 'https://www.sony.com/',
    'lg': 'https://www.lg.com/',
    'intel': 'https://www.intel.com/',
    'amd': 'https://www.amd.com/',
}

CATEGORY_ICONS = {
    'smartphones': '📱',
    'laptops': '💻',
    'tablets': '🖥️',
    'accessories': '🎧',
    'cpu': '⚙️',
    'gpu': '🎮',
    'ram': '💾',
    'ssd': '💿',
    'psu': '🔌',
    'cooling': '🌀',
    'monitors': '🖥️',
    'motherboards': '🔧',
    'resistors': '⚡',
    'capacitors': '🔋',
    'transistors': '🔀',
    'ics': '🔲',
    'diodes': '➡️',
    'inductors': '〰️',
    'connectors': '🔗',
    'relays': '🔄',
}

ENGINEERING_FILTERS = {
    'nominal': ('value', 'resistance', 'capacitance', 'inductance', 'capacity'),
    'power': ('power', 'wattage', 'tdp', 'tdp_rated'),
    'voltage': ('voltage', 'supply_voltage', 'vceo', 'vr', 'vrrm', 'vf', 'coil_voltage'),
    'current': ('current', 'ic', 'id', 'if'),
    'tolerance': ('tolerance',),
    'type': ('type',),
    'pins': ('pins', 'pin_count', 'contact_count'),
    'pitch': ('pitch',),
    'frequency': ('frequency', 'base_clock', 'boost_clock', 'refresh_rate', 'bandwidth'),
    'capacity': (
        'capacity',
        'ram',
        'storage',
        'vram',
        'battery',
        'battery_earbuds',
        'battery_case',
        'cache_l3',
    ),
    'display': ('screen_size', 'resolution', 'panel', 'display', 'hdr'),
    'platform': ('socket', 'chipset', 'chipset_support', 'chip', 'cpu', 'gpu', 'gpu_chip', 'process'),
    'connectivity': (
        'connectivity',
        'network',
        'outputs',
        'inputs',
        'power_conn',
        'm2_slots',
        'pcie',
        'ram_slots',
        'charging',
        'os',
        'codec',
    ),
    # Расширенные фильтры — раньше эти поля показывались в чипах, но не были
    # кликабельны. Теперь юзер может зайти в категорию и отфильтровать товары
    # по тому же интерфейсу / форм-фактору / диэлектрику.
    'form_factor': ('form_factor',),
    'interface': ('interface',),
    'dielectric': ('dielectric',),
    'mounting': ('mounting',),
    'material': ('material', 'contact_material', 'flux_core'),
    'application': ('application',),
    'size': ('length', 'width', 'diameter', 'size', 'board_size', 'hole_count', 'points', 'power_rails'),
    'wire': ('gauge', 'section', 'color'),
    'configuration': ('configuration', 'orientation', 'gender'),
    'temperature_range': (
        'temperature_range',
        'operating_temp',
        'max_temp',
        'temp_max',
        'min_temp',
        'melting_point',
    ),
    'compatibility': ('compatibility',),
    'mode': ('mode', 'signal'),
    'safety': ('safety',),
    # package_type — отдельный фильтр на Product.package_type (CharField),
    # тут только JSON-параметры. Чтобы не было коллизии GET-имени.
}

PROJECT_CART_SESSION_KEY = 'project_cart_context'


def _param_text(product, keys):
    params = product.parameters or {}
    values = []
    for key in keys:
        value = params.get(key)
        if value not in (None, ''):
            values.append(str(value))
    return ' '.join(values)


def _matches_param(product, name, needle):
    if not needle:
        return True
    return needle.lower() in _param_text(product, ENGINEERING_FILTERS[name]).lower()


def _collection_count(items):
    return len(items) if isinstance(items, list) else items.count()


def _official_product_resources(product):
    params = product.parameters or {}
    resources = []
    seen = set()

    def add(label, url, kind):
        if not url or url in seen:
            return
        seen.add(url)
        resources.append({'label': label, 'url': url, 'kind': kind})

    add('Datasheet', product.datasheet_url, 'datasheet')
    for key, label in (
        ('product_url', 'Страница товара'),
        ('manufacturer_url', 'Страница производителя'),
        ('image_source_url', 'Источник фото'),
        ('cad_model_url', 'CAD/3D-модель'),
    ):
        add(label, params.get(key), key)
    add(
        f'Сайт: {product.get_manufacturer_display()}',
        MANUFACTURER_RESOURCE_URLS.get(product.manufacturer),
        'manufacturer',
    )
    return resources


from django.core.cache import cache as _django_cache


def _engineering_filter_options(products, cache_key=None, ttl=300):
    """Опции для engineering-фильтров (nominal/voltage/power/...).

    На каждый запрос вызывается ДВАЖДЫ — в index() и в _catalog_filter_context().
    Раньше материализовался queryset 89-100 раз → ~30 ms PostgreSQL / 80 ms SQLite.
    С кешем 5 мин (TTL=300) ноль на всех вызовах кроме первого.

    `cache_key` — стабильный ключ кеша; None → fallback на обычный compute
    (для случаев когда переданы уже отфильтрованные products).
    """
    if cache_key:
        cached = _django_cache.get(cache_key)
        if cached is not None:
            return cached
        result = build_filter_options(products)
        _django_cache.set(cache_key, result, ttl)
        return result
    return build_filter_options(products)


def _project_context_from_payload(request, payload, added_items, limited, skipped):
    raw_project = payload.get('project') if isinstance(payload.get('project'), dict) else {}
    project_id = raw_project.get('id') or payload.get('project_id')
    project_name = (raw_project.get('name') or payload.get('project_name') or '').strip()
    project_url = ''

    if project_id:
        try:
            project_id = int(project_id)
        except (TypeError, ValueError):
            project_id = None

    if project_id:
        try:
            from Dolg_APP.models import SchematicProject

            project_qs = SchematicProject.objects.filter(pk=project_id)
            if request.user.is_authenticated:
                project_qs = project_qs.filter(Q(user=request.user) | Q(is_demo=True))
            else:
                project_qs = project_qs.filter(is_demo=True)
            project = project_qs.first()
            if project:
                project_name = project.name
                project_url = f'{reverse("hello:simulation")}?project={project.id}'
            else:
                project_id = None
        except Exception:
            project_id = None

    if not project_name:
        project_name = 'Схема без названия'

    warnings = []
    warnings.extend(limited[:5])
    warnings.extend(skipped[:5])

    return {
        'source': (payload.get('source') or raw_project.get('source') or 'BOM').strip()[:40],
        'project_id': project_id,
        'project_name': project_name[:160],
        'project_url': project_url,
        'created_at': timezone.now().isoformat(),
        'total_positions': len(added_items),
        'total_quantity': sum(item['quantity'] for item in added_items),
        'grand_total': round(sum(item['line_total'] for item in added_items), 2),
        'items': added_items[:12],
        'warnings': warnings[:10],
    }


def _apply_filters(products, request):
    """Apply GET-parameter filters shared by index and category views.

    Поиск умный (`smart_search`): multi-token парсинг + fuzzy fallback на typos.
    Старый icontains-only оставлен как 1-токенный case через smart_search().
    """
    return apply_catalog_filters(products, request.GET)


SORT_OPTIONS = [
    ('', 'По релевантности'),
    ('price_asc', '💰 Цена ↑ (дешевле)'),
    ('price_desc', '💰 Цена ↓ (дороже)'),
    ('rating', '⭐ По рейтингу'),
    ('newest', '🆕 Сначала новые'),
    ('stock', '📦 По наличию'),
    ('name', '🔤 По названию'),
]


def _apply_sort(products, sort_key):
    """Sort products list by GET-parameter `?sort=`.

    Принимает list (apply_catalog_filters уже материализует QuerySet).
    Возвращает новый отсортированный список — не мутирует входной.
    """
    if not sort_key:
        return products
    if sort_key == 'price_asc':
        return sorted(products, key=lambda p: float(p.price))
    if sort_key == 'price_desc':
        return sorted(products, key=lambda p: float(p.price), reverse=True)
    if sort_key == 'newest':
        return sorted(products, key=lambda p: p.created_at, reverse=True)
    if sort_key == 'stock':
        return sorted(products, key=lambda p: p.stock, reverse=True)
    if sort_key == 'name':
        return sorted(products, key=lambda p: (p.name or '').lower())
    if sort_key == 'rating':
        from shop.card_helpers import product_rating

        return sorted(products, key=lambda p: product_rating(p)['value'], reverse=True)
    return products


def _catalog_filter_context(request, base_url, filtered_products, manufacturer_choices, lifecycle_choices):
    active = clean_params(request.GET)
    pagination_query = querystring_with(request.GET, 'page', '')
    return {
        'active_filters': build_active_filter_tags(
            request.GET,
            base_url,
            manufacturer_choices=manufacturer_choices,
            lifecycle_choices=lifecycle_choices,
        ),
        'active_filter_keys': set(active.keys()),
        'hidden_filter_inputs': hidden_filter_inputs(request.GET),
        'pagination_query': pagination_query,
        'facets': compute_catalog_facets(filtered_products),
        'active_price_min': active.get('price_min', ''),
        'active_price_max': active.get('price_max', ''),
        'active_part_number': active.get('part_number', ''),
        'active_hide_eol': active.get('hide_eol', ''),
        'active_has_datasheet': active.get('has_datasheet', ''),
        'active_has_spice_model': active.get('has_spice_model', ''),
        'active_has_cad_model': active.get('has_cad_model', ''),
        'active_form_factor': active.get('form_factor', ''),
        'active_interface': active.get('interface', ''),
        'active_dielectric': active.get('dielectric', ''),
        'active_mounting': active.get('mounting', ''),
        'catalog_filter_names': CATALOG_FILTERS,
    }


def healthz(request):
    """Лёгкий liveness-probe для Docker/Kubernetes healthcheck. Не лезет в
    БД (не падает на медленной БД при ConnectionPool exhaustion) — нужен
    отдельный /readyz/ для readiness, если когда-нибудь понадобится."""
    return HttpResponse('ok', content_type='text/plain')


def readyz(request):
    """Readiness probe: confirms that core DB-backed modules are usable."""
    try:
        payload = {
            'ok': True,
            'products': Product.objects.count(),
            'categories': Category.objects.count(),
        }
        try:
            from knowledge.models import Article

            payload['articles'] = Article.objects.filter(is_published=True).count()
        except Exception as exc:  # pragma: no cover - defensive probe
            payload['ok'] = False
            payload['knowledge_error'] = str(exc)
        try:
            from Dolg_APP.models import SchematicProject

            payload['demo_projects'] = SchematicProject.objects.filter(is_demo=True).count()
        except Exception as exc:  # pragma: no cover - defensive probe
            payload['ok'] = False
            payload['projects_error'] = str(exc)
        return JsonResponse(payload, status=200 if payload.get('ok') else 503)
    except Exception as exc:  # pragma: no cover - readiness endpoint must never explode
        return JsonResponse({'ok': False, 'error': str(exc)}, status=503)


def about(request):
    """Страница «О сайте» — что это за платформа, какие модули и кому полезна.

    Цифры (товары/статьи/проекты) подтягиваются из БД, чтобы не протухали.
    Раньше в шаблоне были hardcode «26+ компонентов, 9+ статей, 17+ тестов» —
    после bulk-import'ов это уже не отражает реальность.
    """
    from Dolg_APP.models import SchematicProject
    from knowledge.models import Article

    stats = {
        'products': Product.objects.count(),
        'reb_products': Product.objects.filter(category__slug__in=REB_SLUGS).count(),
        'categories': Category.objects.count(),
        'demo_projects': SchematicProject.objects.filter(is_demo=True).count(),
        'articles': Article.objects.count(),
        # Тесты считаются на момент последнего прогона CI. Если есть
        # docs/TESTS_AND_REPORTS — оттуда. Пока — статичная цифра, легче
        # пересчитать через скрипт `manage.py refresh_about_stats`.
        'tests': 263,
    }
    return render(request, 'shop/about.html', {'stats': stats})


def _tool_results(query):
    tools = [
        {
            'title': 'Engineering Review',
            'description': 'Design Health Score, DRC/ERC, BOM risk, derating, diagnostics, measurements and PDF report.',
            'url': reverse('hello:projects'),
            'keywords': 'review engineering health score диагностика перегрев запас derating fault faults ошибки gnd источник bom risk pdf отчет',
        },
        {
            'title': 'Expert System Review',
            'description': 'Rule packs, unit-safe номиналы, Z3 constraint подбор, Lark CAD parsing and fuzzy risk scoring before neural analysis.',
            'url': reverse('hello:projects'),
            'keywords': 'expert system rule-engine jsonschema pint units unit-safe z3 constraint solver lark parser scikit-fuzzy fuzzy risk экспертная система правила номиналы подбор ограничений риск',
        },
        {
            'title': 'CAD Import to Review',
            'description': 'Import LTspice/SPICE netlist and KiCad subset, normalize scheme_data, then run DOLG review.',
            'url': reverse('hello:cad'),
            'keywords': 'import cad ltspice kicad spice netlist asc sch импорт схема нормализация review',
        },
        {
            'title': 'CAD редактор',
            'description': '2D САПР, semantic pins/nets, DRC, BOM, A3-экспорт и связь с каталогом.',
            'url': reverse('hello:cad'),
            'keywords': 'cad сапр чертеж схема drc bom гост a3 wiring',
        },
        {
            'title': 'Симуляция схем',
            'description': 'Редактор принципиальных схем, ngspice.wasm, графики, BOM и экспорт.',
            'url': reverse('hello:simulation'),
            'keywords': 'симуляция spice ngspice схема график расчет анализ',
        },
        {
            'title': 'Pro-аналитика симуляции',
            'description': 'FFT-спектр, Bode plot, THD/SINAD/ENOB, parameter sweep, Monte Carlo tolerance, server-side fallback и сохранение измерений проекта.',
            'url': reverse('hello:simulation'),
            'keywords': 'pro fft spectrum bode plot monte carlo tolerance scipy numpy matplotlib pandas fallback solver mna thd sinad enob signal quality parameter sweep what-if sweep спектр ачх фчх искажения качество сигнала гармоники допуск измерение projectmeasurement',
        },
        {
            'title': 'Инженерная лаборатория',
            'description': 'Расчеты и оценки для транзисторного ключа, NE555, стабилизатора, RC и теплового запаса.',
            'url': reverse('knowledge:engineering_lab'),
            'keywords': 'лаборатория расчеты измерения транзисторный ключ ne555 стабилизатор тепло rms duty cycle',
        },
        {
            'title': 'Практикум обучения',
            'description': 'Уроки с расчетами, сборкой схем и измерениями результата симуляции.',
            'url': reverse('knowledge:learning_index'),
            'keywords': 'обучение уроки практикум задания закон ома делитель rc фильтр измерение транзистор ne555 стабилизатор',
        },
        {
            'title': 'Проекты',
            'description': 'Личный список схем, демо-проекты, версии и история запусков.',
            'url': reverse('hello:projects'),
            'keywords': 'проекты схемы версии история демонстрация',
        },
        {
            'title': 'Сравнение товаров',
            'description': 'Сравнение параметров компонентов и электроники с подсветкой отличий.',
            'url': reverse('shop:compare'),
            'keywords': 'сравнение аналоги параметры товары',
        },
        {
            'title': 'Корзина и BOM',
            'description': 'Переход от спецификации схемы к заказу компонентов.',
            'url': reverse('shop:cart'),
            'keywords': 'корзина bom заказ спецификация компоненты',
        },
    ]
    q = (query or '').lower()
    if not q:
        return tools[:4]
    return [
        item
        for item in tools
        if q in item['title'].lower() or q in item['description'].lower() or q in item['keywords'].lower()
    ]


def _project_results_for_user(user, query):
    try:
        from Dolg_APP.models import SchematicProject
    except Exception:
        return []

    base = SchematicProject.objects.select_related('user')
    if user.is_authenticated:
        base = base.filter(Q(user=user) | Q(is_demo=True))
    else:
        base = base.filter(is_demo=True)
    if query:
        base = base.filter(Q(name__icontains=query) | Q(description__icontains=query))
    return base.order_by('-is_demo', '-updated_at')[:8]


def _legal_source_results(query, *, limit=8):
    if not query:
        return []
    try:
        from knowledge.models import Article, ArticleMaterial
        from knowledge.services.legal_sources import find_legal_sources

        overview = Article.objects.filter(
            slug='otkrytye-istochniki-i-dokumentatsiya-dolg',
            is_published=True,
        ).first()
        by_url = {}
        for item in find_legal_sources(query, limit=limit):
            material = (
                ArticleMaterial.objects.select_related('article', 'article__category')
                .filter(is_public=True, url=item['url'], article__is_published=True)
                .order_by('article__order', 'order')
                .first()
            )
            article = material.article if material else overview
            by_url[item['url']] = {
                'id': item['id'],
                'title': item['title'],
                'description': item.get('description', ''),
                'topic': item.get('topic', ''),
                'url': item['url'],
                'article': article,
                'article_url': reverse('knowledge:article', args=[article.slug]) if article else item['url'],
                'kicker': f'Источник · {item.get("topic", "")}',
            }

        materials = (
            ArticleMaterial.objects.select_related('article', 'article__category')
            .filter(is_public=True, article__is_published=True)
            .filter(Q(title__icontains=query) | Q(description__icontains=query) | Q(url__icontains=query))
            .order_by('article__order', 'order')[:limit]
        )
        for material in materials:
            if not material.url or material.url in by_url:
                continue
            by_url[material.url] = {
                'id': f'material-{material.id}',
                'title': material.title.replace('Источник: ', ''),
                'description': material.description,
                'topic': material.material_type,
                'url': material.url,
                'article': material.article,
                'article_url': reverse('knowledge:article', args=[material.article.slug]),
                'kicker': f'Материал · {material.article.title}',
            }
        return list(by_url.values())[:limit]
    except Exception:
        return []


def global_search(request):
    query = (request.GET.get('q') or '').strip()

    products = Product.objects.select_related('category').none()
    categories = Category.objects.none()
    articles = []
    learning_lessons = []
    projects = []
    source_results = []

    if query:
        products = (
            Product.objects.select_related('category')
            .filter(
                Q(name__icontains=query)
                | Q(description__icontains=query)
                | Q(part_number__icontains=query)
                | Q(package_type__icontains=query)
                | Q(manufacturer__icontains=query)
                | Q(category__name__icontains=query)
            )
            .order_by('category__name', 'name')[:12]
        )
        categories = Category.objects.filter(
            Q(name__icontains=query) | Q(description__icontains=query) | Q(slug__icontains=query)
        ).order_by('name')[:8]

        try:
            from knowledge.models import Article

            articles = (
                Article.objects.select_related('category')
                .filter(is_published=True)
                .filter(
                    Q(title__icontains=query)
                    | Q(summary__icontains=query)
                    | Q(body__icontains=query)
                    | Q(related_components_note__icontains=query)
                    | Q(category__name__icontains=query)
                )
                .order_by('category__order', 'order', 'title')[:10]
            )
        except Exception:
            articles = []

        try:
            from knowledge.models import LearningLesson

            query_variants = {query, query.lower(), query.upper(), query.capitalize(), query.title()}
            learning_filter = Q()
            for term in query_variants:
                learning_filter |= (
                    Q(title__icontains=term)
                    | Q(summary__icontains=term)
                    | Q(theory__icontains=term)
                    | Q(formula__icontains=term)
                    | Q(track__title__icontains=term)
                    | Q(tasks__title__icontains=term)
                    | Q(tasks__prompt__icontains=term)
                )
            learning_lessons = (
                LearningLesson.objects.select_related('track')
                .filter(is_published=True, track__is_published=True)
                .filter(learning_filter)
                .distinct()
                .order_by('track__order', 'order', 'title')[:8]
            )
        except Exception:
            learning_lessons = []

        projects = _project_results_for_user(request.user, query)
        source_results = _legal_source_results(query)
    else:
        products = Product.objects.select_related('category').order_by('-created_at')[:8]
        try:
            from knowledge.models import Article

            articles = (
                Article.objects.select_related('category')
                .filter(is_published=True)
                .order_by('category__order', 'order')[:6]
            )
        except Exception:
            articles = []
        try:
            from knowledge.models import LearningLesson

            learning_lessons = (
                LearningLesson.objects.select_related('track')
                .filter(is_published=True, track__is_published=True)
                .order_by('track__order', 'order', 'title')[:6]
            )
        except Exception:
            learning_lessons = []
        projects = _project_results_for_user(request.user, '')

    tools = _tool_results(query)
    total_results = (
        len(products)
        + len(categories)
        + len(articles)
        + len(learning_lessons)
        + len(projects)
        + len(tools)
        + len(source_results)
    )
    return render(
        request,
        'shop/search_results.html',
        {
            'query': query,
            'products': products,
            'categories_found': categories,
            'articles': articles,
            'learning_lessons': learning_lessons,
            'projects': projects,
            'tools': tools,
            'source_results': source_results,
            'total_results': total_results,
        },
    )


def demo_route(request):
    featured_product = (
        Product.objects.select_related('category')
        .filter(category__slug__in=REB_SLUGS, image__isnull=False)
        .order_by('category__slug', 'price')
        .first()
    )

    featured_article = None
    try:
        from knowledge.models import Article

        featured_article = (
            Article.objects.select_related('category')
            .filter(is_published=True)
            .order_by('category__order', 'order', 'title')
            .first()
        )
    except Exception:
        featured_article = None

    demo_project = None
    try:
        from Dolg_APP.models import SchematicProject

        demo_project = SchematicProject.objects.filter(is_demo=True).order_by('-difficulty', 'name').first()
    except Exception:
        demo_project = None

    steps = [
        {
            'label': 'Каталог',
            'title': 'Найти реальный компонент',
            'description': 'Открываем товар с фото, параметрами, lifecycle и datasheet.',
            'url': featured_product.get_absolute_url()
            if hasattr(featured_product, 'get_absolute_url')
            else (
                reverse('shop:product_detail', args=[featured_product.slug])
                if featured_product
                else reverse('shop:index')
            ),
        },
        {
            'label': 'Энциклопедия',
            'title': 'Понять применение',
            'description': 'Переходим к статье, материалам, схемам включения и практическим подсказкам.',
            'url': reverse('knowledge:article', args=[featured_article.slug])
            if featured_article
            else reverse('knowledge:index'),
        },
        {
            'label': 'CAD',
            'title': 'Собрать чертеж',
            'description': 'Используем semantic pins/nets, smart wiring, DRC и спецификацию.',
            'url': reverse('hello:cad'),
        },
        {
            'label': 'Симуляция',
            'title': 'Проверить поведение',
            'description': 'Запускаем расчет, смотрим графики, предупреждения и результат по узлам.',
            'url': reverse('hello:simulation'),
        },
        {
            'label': 'BOM',
            'title': 'Собрать спецификацию',
            'description': 'Группируем компоненты, проверяем наличие, аналоги и предупреждения.',
            'url': reverse('hello:simulation'),
        },
        {
            'label': 'Заказ',
            'title': 'Перейти к корзине',
            'description': 'Добавляем позиции BOM в корзину и оформляем учебный заказ.',
            'url': reverse('shop:cart'),
        },
    ]
    return render(
        request,
        'shop/demo_route.html',
        {
            'steps': steps,
            'featured_product': featured_product,
            'featured_article': featured_article,
            'demo_project': demo_project,
        },
    )


def index(request):
    # ОПТИМИЗАЦИЯ (2026-05-19): аннотируем .products_count_cached одним SQL,
    # шаблон index.html использовал {{ cat.products.count }} в loop = 12+
    # запросов на каждом рендере главной. Фильтрация reb/consumer теперь
    # в Python (1 раз материализуем, потом фильтруем по slug).
    all_categories_list = list(Category.objects.annotate(products_count_cached=Count('products')).all())
    reb_categories = [c for c in all_categories_list if c.slug in REB_SLUGS]
    consumer_categories = [c for c in all_categories_list if c.slug not in REB_SLUGS]

    products = Product.objects.select_related('category').all()
    # Кеш на 5 мин — engineering-options меняются только при добавлении/удалении
    # товаров (редко в проде). Ключ инвалидируется через ProductCount-стемп.
    options_stamp = Product.objects.count()
    param_filter_options = _engineering_filter_options(
        products, cache_key=f'eng_filter_opts:catalog:{options_stamp}'
    )
    products, query = _apply_filters(products, request)
    products = _apply_sort(products, request.GET.get('sort', ''))
    product_total = _collection_count(products)
    has_active_filters = bool(clean_params(request.GET))

    # Sidebar filter options (unique values from current queryset)
    manufacturers = Product.objects.values_list('manufacturer', flat=True).distinct()
    packages = Product.objects.exclude(package_type='').values_list('package_type', flat=True).distinct()
    # Facet-counts по уже-отфильтрованному набору товаров — для UI sidebar.
    # Показываем «Vishay (8) · Yageo (4)» вместо плоского списка фильтров.
    catalog_shelves = []
    if not has_active_filters:
        for category in all_categories_list:
            shelf_products = list(
                Product.objects.select_related('category')
                .filter(category=category)
                .order_by('-created_at', 'name')[:6]
            )
            if not shelf_products:
                continue
            catalog_shelves.append(
                {
                    'category': category,
                    'products': shelf_products,
                    # Берём аннотацию вместо отдельного .count() — экономия 1 SQL на категорию.
                    'count': category.products_count_cached,
                }
            )

    # Без фильтров «Новые позиции» обязаны быть disjoint с полками категорий —
    # иначе те же товары рендерятся дважды (полка → ниже общий список).
    if catalog_shelves:
        shelved_ids = {p.id for shelf in catalog_shelves for p in shelf['products']}
        if shelved_ids:
            products = [product for product in products if product.id not in shelved_ids]

    # Пагинация: на проде каталог может вырасти за 1000 — без paginator
    # страница тащит весь queryset (72 сейчас, но защищаемся на будущее).
    paginator = Paginator(products, PRODUCTS_PER_PAGE)
    page_obj = paginator.get_page(request.GET.get('page'))

    manufacturer_choices = dict(Product.MANUFACTURER_CHOICES)
    lifecycle_choices = dict(Product.LIFECYCLE_CHOICES)
    manufacturer_choices = dict(Product.MANUFACTURER_CHOICES)
    lifecycle_choices = dict(Product.LIFECYCLE_CHOICES)
    context = {
        'products': page_obj,  # iterable — шаблон не трогаем
        'page_obj': page_obj,  # для ссылок previous/next
        'paginator': paginator,
        'product_total': product_total,
        'catalog_shelves': catalog_shelves,
        'has_active_filters': has_active_filters,
        'categories': all_categories_list,
        'reb_categories': reb_categories,
        'consumer_categories': consumer_categories,
        'query': query,
        'manufacturers': sorted(set(manufacturers)),
        'packages': sorted(set(packages)),
        'facets': compute_catalog_facets(products),
        'active_manufacturer': request.GET.get('manufacturer', ''),
        'active_lifecycle': request.GET.get('lifecycle', ''),
        'active_package': request.GET.get('package', ''),
        'active_in_stock': request.GET.get('in_stock', ''),
        'active_nominal': request.GET.get('nominal', ''),
        'active_power': request.GET.get('power', ''),
        'active_voltage': request.GET.get('voltage', ''),
        'active_tolerance': request.GET.get('tolerance', ''),
        'active_sort': request.GET.get('sort', ''),
        'sort_options': SORT_OPTIONS,
        'param_filter_options': param_filter_options,
        'manufacturer_choices': manufacturer_choices,
        'lifecycle_choices': lifecycle_choices,
        'category_icons': CATEGORY_ICONS,
    }
    context.update(
        _catalog_filter_context(
            request, reverse('shop:index'), products, manufacturer_choices, lifecycle_choices
        )
    )
    return render(request, 'shop/index.html', context)


def category_products(request, slug):
    category = get_object_or_404(Category, slug=slug)
    all_categories = Category.objects.all()
    reb_categories = all_categories.filter(slug__in=REB_SLUGS)
    consumer_categories = all_categories.exclude(slug__in=REB_SLUGS)

    products = category.products.all()
    options_stamp = category.products.count()
    param_filter_options = _engineering_filter_options(
        products, cache_key=f'eng_filter_opts:cat:{category.slug}:{options_stamp}'
    )
    products, query = _apply_filters(products, request)
    products = _apply_sort(products, request.GET.get('sort', ''))

    manufacturers = category.products.values_list('manufacturer', flat=True).distinct()
    packages = category.products.exclude(package_type='').values_list('package_type', flat=True).distinct()
    paginator = Paginator(products, PRODUCTS_PER_PAGE)
    page_obj = paginator.get_page(request.GET.get('page'))
    manufacturer_choices = dict(Product.MANUFACTURER_CHOICES)
    lifecycle_choices = dict(Product.LIFECYCLE_CHOICES)

    context = {
        'products': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'categories': all_categories,
        'reb_categories': reb_categories,
        'consumer_categories': consumer_categories,
        'current_category': category,
        'is_reb': category.slug in REB_SLUGS,
        'query': query,
        'manufacturers': sorted(set(manufacturers)),
        'packages': sorted(set(packages)),
        'facets': compute_catalog_facets(products),
        'active_manufacturer': request.GET.get('manufacturer', ''),
        'active_lifecycle': request.GET.get('lifecycle', ''),
        'active_package': request.GET.get('package', ''),
        'active_in_stock': request.GET.get('in_stock', ''),
        'active_nominal': request.GET.get('nominal', ''),
        'active_power': request.GET.get('power', ''),
        'active_voltage': request.GET.get('voltage', ''),
        'active_tolerance': request.GET.get('tolerance', ''),
        'active_sort': request.GET.get('sort', ''),
        'sort_options': SORT_OPTIONS,
        'param_filter_options': param_filter_options,
        'manufacturer_choices': manufacturer_choices,
        'lifecycle_choices': lifecycle_choices,
        'category_icons': CATEGORY_ICONS,
    }
    context.update(
        _catalog_filter_context(
            request,
            reverse('shop:category', args=[category.slug]),
            products,
            manufacturer_choices,
            lifecycle_choices,
        )
    )
    return render(request, 'shop/category.html', context)


RECENTLY_VIEWED_LIMIT = 5
RECENTLY_VIEWED_KEY = '_recently_viewed_pids'


def product_detail(request, slug):
    product = get_object_or_404(Product, slug=slug)
    # Recently-viewed: для auth-юзеров пишем в персистентную таблицу
    # ViewedProduct (retention 90 дней, очистка через cron). Для guest
    # остаётся сессионная история (см. ниже recently_viewed_ids).
    if request.user.is_authenticated:
        from .models import ViewedProduct

        ViewedProduct.objects.update_or_create(
            user=request.user,
            product=product,
        )
    related_products = Product.objects.filter(category=product.category).exclude(id=product.id)[:4]
    alternative_products = (
        Product.objects.filter(
            category=product.category, stock__gt=0, lifecycle_status__in=['active', 'nrnd']
        )
        .exclude(id=product.id)
        .order_by('price', 'name')[:6]
    )

    engineering_warnings = []
    if product.stock <= 0:
        engineering_warnings.append('Товара нет на складе: для BOM лучше выбрать аналог в наличии.')
    elif product.stock < 10:
        engineering_warnings.append(f'Небольшой остаток на складе: {product.stock} шт.')
    if product.lifecycle_status != 'active':
        engineering_warnings.append(
            f'Lifecycle: {product.get_lifecycle_status_display()}. Перед заказом стоит проверить замену.'
        )
    if product.is_reb() and not product.datasheet_url:
        engineering_warnings.append(
            'Нет datasheet-ссылки: проверьте корпус, распиновку и предельные параметры вручную.'
        )
    if product.is_reb() and not (product.parameters or {}).get('spice_model'):
        engineering_warnings.append(
            'SPICE-модель не указана: симулятор использует типовую модель или упрощенный расчет.'
        )

    template_map = {
        'resistors': [
            (
                'Делитель напряжения',
                'Подбор плеч делителя и проверка мощности резисторов.',
                '/simulation/?tour=1',
            ),
            ('Измерительный шунт', 'Оценка падения напряжения и рассеиваемой мощности.', '/cad/'),
        ],
        'capacitors': [
            (
                'RC-фильтр',
                'Частота среза, переходный процесс и подбор рабочего напряжения.',
                '/simulation/?tour=1',
            ),
            ('Фильтр питания', 'Сглаживание пульсаций и проверка ESR/температуры.', '/cad/'),
        ],
        'diodes': [
            (
                'Выпрямитель',
                'Проверка падения напряжения, обратного напряжения и нагрева.',
                '/simulation/?tour=1',
            ),
            ('Защита входа', 'TVS/сигнальные диоды в цепях защиты.', '/cad/'),
        ],
        'transistors': [
            ('Транзисторный ключ', 'Расчет базового/затворного управления и потерь.', '/simulation/?tour=1'),
            ('Усилительный каскад', 'Рабочая точка и малосигнальный режим.', '/cad/'),
        ],
        'ics': [
            (
                'Стабилизатор или таймер',
                'Проверка питания, обвязки и типовой схемы включения.',
                '/simulation/?tour=1',
            ),
            ('ОУ/логика', 'Питание, входы, выходы и неподключенные пины.', '/cad/'),
        ],
        'inductors': [
            ('LC-фильтр', 'Резонанс, ток насыщения и сопротивление обмотки.', '/simulation/?tour=1'),
        ],
        'connectors': [
            ('Ввод питания/сигнала', 'Проверка контактов, шага и допустимого тока.', '/cad/'),
        ],
        'relays': [
            (
                'Коммутация нагрузки',
                'Катушка, контакты, защитный диод и ток нагрузки.',
                '/simulation/?tour=1',
            ),
        ],
    }
    compatible_templates = template_map.get(
        product.category.slug,
        [
            (
                'Применение в проекте',
                'Добавьте товар в проектную корзину или используйте как позицию спецификации.',
                reverse('hello:cad'),
            ),
        ],
    )

    related_articles = []
    try:
        from knowledge.models import Article

        terms = [product.category.name, product.category.slug, product.part_number, product.name]
        article_query = Q()
        for term in terms:
            if term:
                article_query |= (
                    Q(title__icontains=term)
                    | Q(summary__icontains=term)
                    | Q(body__icontains=term)
                    | Q(related_components_note__icontains=term)
                )
        related_articles = (
            Article.objects.select_related('category')
            .filter(is_published=True)
            .filter(article_query)
            .order_by('category__order', 'order', 'title')[:6]
        )
    except Exception:
        related_articles = []

    # Recently viewed: храним последние RECENTLY_VIEWED_LIMIT product-id-ов в
    # session. Текущий товар двигаем в начало (или добавляем). Сама панель
    # рендерится в product_detail.html по списку recently_viewed (без id-текущего).
    history = request.session.get(RECENTLY_VIEWED_KEY, []) or []
    history = [pid for pid in history if pid != product.id]
    history.insert(0, product.id)
    request.session[RECENTLY_VIEWED_KEY] = history[:RECENTLY_VIEWED_LIMIT]
    # Загружаем продукты в порядке истории (исключая текущий, чтобы не дублировать).
    other_ids = [pid for pid in history if pid != product.id][:RECENTLY_VIEWED_LIMIT]
    rv_qs = Product.objects.filter(id__in=other_ids)
    rv_map = {p.id: p for p in rv_qs}
    recently_viewed = [rv_map[pid] for pid in other_ids if pid in rv_map]

    context = {
        'product': product,
        'related_products': related_products,
        'alternative_products': alternative_products,
        'engineering_warnings': engineering_warnings,
        'compatible_templates': compatible_templates,
        'related_articles': related_articles,
        'recently_viewed': recently_viewed,
        'official_resources': _official_product_resources(product),
    }
    return render(request, 'shop/product_detail.html', context)


@require_POST
def add_to_cart(request, slug):
    from django.contrib import messages as _msg

    product = get_object_or_404(Product, slug=slug)
    cart_filter = _cart_owner_filter(request)
    try:
        quantity = max(1, int(request.POST.get('quantity', 1)))
    except (ValueError, TypeError):
        quantity = 1

    if product.stock <= 0:
        _msg.error(request, f'Товара «{product.name}» нет на складе')
        return redirect('shop:product_detail', slug=slug)

    # Сколько уже лежит в корзине (если до этого добавляли тот же товар).
    existing = CartItem.objects.filter(product=product, **cart_filter).first()
    already_in_cart = existing.quantity if existing else 0
    requested_total = already_in_cart + quantity
    # Не даём суммарно превысить остаток — клампим к product.stock.
    if requested_total > product.stock:
        capped_quantity = product.stock - already_in_cart
        if capped_quantity <= 0:
            _msg.warning(
                request,
                f'В корзине уже {already_in_cart} шт. товара «{product.name}» — '
                f'это весь доступный остаток ({product.stock}).',
            )
            return redirect('shop:cart')
        _msg.warning(
            request,
            f'Доступно только {product.stock} шт. товара «{product.name}». '
            f'Добавлено {capped_quantity} (а не {quantity}).',
        )
        quantity = capped_quantity

    cart_item, created = CartItem.objects.get_or_create(
        product=product,
        defaults={'quantity': quantity, **_cart_owner_create_defaults(request)},
        **cart_filter,
    )
    if not created:
        cart_item.quantity = min(cart_item.quantity + quantity, product.stock)
        cart_item.save(update_fields=['quantity'])

    # Guest: явно говорим, что корзина живёт ограниченно (сессия). Без этого
    # юзер может думать что заказ оформился — а он висит в session-CartItem.
    if not request.user.is_authenticated:
        _msg.info(
            request,
            f'✓ «{product.name}» добавлен в корзину. Корзина сохранится в текущей сессии — '
            f'войдите, чтобы оформить заказ.',
        )
    else:
        _msg.success(request, f'✓ «{product.name}» добавлен в корзину')
    return redirect('shop:cart')


def cart(request):
    cart_items = CartItem.objects.select_related('product').filter(**_cart_owner_filter(request))

    total_price = sum(item.get_total_price() for item in cart_items)

    context = {
        'cart_items': cart_items,
        'total_price': total_price,
        'project_cart_context': request.session.get(PROJECT_CART_SESSION_KEY),
    }
    return render(request, 'shop/cart.html', context)


@require_POST
def clear_project_cart_context(request):
    request.session.pop(PROJECT_CART_SESSION_KEY, None)
    request.session.modified = True
    return redirect('shop:cart')


@require_POST
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, **_cart_owner_filter(request))
    cart_item.delete()
    return redirect('shop:cart')


@require_POST
def update_cart_item(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, **_cart_owner_filter(request))
    try:
        quantity = int(request.POST.get('quantity', 1))
    except (ValueError, TypeError):
        quantity = 1

    if quantity > 0 and cart_item.product.stock > 0:
        cart_item.quantity = min(quantity, cart_item.product.stock)
        cart_item.save()
    else:
        cart_item.delete()

    return redirect('shop:cart')


# ---------------------------------------------------------------------------
# BOM (Bill of Materials): сопоставление компонентов схемы с товарами каталога
# ---------------------------------------------------------------------------

# Тип компонента в редакторе схем → slug категории в каталоге.
# LED — это диод, поэтому маппится в "diodes".
COMPONENT_TO_CATEGORY = {
    'resistor': 'resistors',
    'capacitor': 'capacitors',
    'inductor': 'inductors',
    'diode': 'diodes',
    'led': 'diodes',
    'transistor': 'transistors',
    'npn': 'transistors',
    'pnp': 'transistors',
    'ic': 'ics',
    'connector': 'connectors',
    'relay': 'relays',
    'switch': 'relays',
}

# Человекочитаемые названия для UI BOM
COMPONENT_LABELS = {
    'resistor': 'Резистор',
    'capacitor': 'Конденсатор',
    'inductor': 'Катушка индуктивности',
    'diode': 'Диод',
    'led': 'Светодиод',
    'transistor': 'Транзистор',
    'ic': 'Микросхема',
    'connector': 'Разъём',
    'relay': 'Реле',
    'switch': 'Переключатель',
    'battery': 'Батарея',
    'ground': 'Земля',
    'node': 'Узел',
}


def _bom_product_dict(p):
    return {
        'id': p.id,
        'slug': p.slug,
        'name': p.name,
        'part_number': p.part_number,
        'price': float(p.price),
        'stock': p.stock,
        'manufacturer': p.get_manufacturer_display(),
        'manufacturer_key': p.manufacturer,
        'category': p.category.name,
        'category_slug': p.category.slug,
        'package_type': p.package_type,
        'lifecycle_status': p.lifecycle_status,
        'lifecycle_display': p.get_lifecycle_status_display(),
        'datasheet_url': p.datasheet_url,
        'parameters': p.parameters or {},
        'url': f'/product/{p.slug}/',
    }


BOM_MAX_COMPONENTS = 1000  # Защита от DoS: вызов с миллионом компонентов
# съест RAM на dict-агрегации и запросах к БД.
# Большая реальная схема — < 500 компонентов.


def _build_bom_matches(components):
    if not isinstance(components, list):
        raise ValueError('components must be a list')
    if len(components) > BOM_MAX_COMPONENTS:
        raise ValueError(f'too many components ({len(components)}); limit is {BOM_MAX_COMPONENTS}')

    explicit_counts = {}
    explicit_components = {}
    counts = {}
    type_components = {}
    for c in components:
        if not isinstance(c, dict):
            continue
        catalog_ref = (c.get('catalog_slug') or c.get('catalog_ref') or c.get('part_number') or '').strip()
        if catalog_ref:
            explicit_counts[catalog_ref] = explicit_counts.get(catalog_ref, 0) + 1
            explicit_components.setdefault(catalog_ref, []).append(c)
            continue
        ctype = (c.get('type') or '').strip().lower()
        if not ctype:
            continue
        counts[ctype] = counts.get(ctype, 0) + 1
        type_components.setdefault(ctype, []).append(c)

    matches = []
    grand_total = 0.0

    for catalog_ref, count in explicit_counts.items():
        product = (
            Product.objects.select_related('category')
            .filter(Q(slug__iexact=catalog_ref) | Q(part_number__iexact=catalog_ref))
            .first()
        )
        entry = {
            'type': 'catalog',
            'label': catalog_ref,
            'count': count,
            'category_slug': product.category.slug if product else None,
            'product': None,
            'alternatives': [],
            'line_total': 0.0,
            'warnings': [],
        }
        if product:
            entry['label'] = product.part_number or product.name
            entry['product'] = _bom_product_dict(product)
            entry['line_total'] = round(float(product.price) * count, 2)
            grand_total += entry['line_total']
            alternatives = (
                Product.objects.select_related('category')
                .filter(category=product.category, stock__gt=0, lifecycle_status__in=['active', 'nrnd'])
                .order_by('price')[:5]
            )
            entry['alternatives'] = [_bom_product_dict(p) for p in alternatives]
            warnings = []
            for component in explicit_components.get(catalog_ref, []):
                warnings.append(nominal_mismatch_warning(component, product))
                warnings.append(missing_spice_model_warning(component, product))
            entry['warnings'] = unique_warnings(warnings)
        else:
            entry['warnings'] = [f'товар каталога "{catalog_ref}" не найден']
        matches.append(entry)

    for ctype, count in counts.items():
        label = COMPONENT_LABELS.get(ctype, ctype.capitalize())
        slug = COMPONENT_TO_CATEGORY.get(ctype)
        entry = {
            'type': ctype,
            'label': label,
            'count': count,
            'category_slug': slug,
            'product': None,
            'alternatives': [],
            'line_total': 0.0,
            'warnings': [],
        }

        if slug:
            products = list(
                Product.objects.select_related('category')
                .filter(category__slug=slug, stock__gt=0, lifecycle_status__in=['active', 'nrnd'])
                .order_by('price')[:5]
            )
            if products:
                best = products[0]
                entry['product'] = _bom_product_dict(best)
                entry['line_total'] = round(float(best.price) * count, 2)
                grand_total += entry['line_total']
                entry['alternatives'] = [_bom_product_dict(p) for p in products]
                warnings = []
                for component in type_components.get(ctype, []):
                    warnings.append(nominal_mismatch_warning(component, best))
                    warnings.append(missing_spice_model_warning(component, best))
                entry['warnings'] = unique_warnings(warnings)

        matches.append(entry)

    matches.sort(key=lambda m: (m['product'] is None, m['label']))
    return matches, round(grand_total, 2), sum(counts.values()) + sum(explicit_counts.values())


@require_GET
def api_component_search(request):
    """Search catalog products for a schematic component properties panel."""
    query = (request.GET.get('q') or '').strip()
    component_type = (request.GET.get('type') or '').strip().lower()
    category_slug = COMPONENT_TO_CATEGORY.get(component_type)

    products = Product.objects.select_related('category').filter(
        stock__gt=0,
        lifecycle_status__in=['active', 'nrnd'],
    )
    if category_slug:
        products = products.filter(category__slug=category_slug)
    if query:
        products = products.filter(
            Q(name__icontains=query)
            | Q(part_number__icontains=query)
            | Q(description__icontains=query)
            | Q(package_type__icontains=query)
        )
    elif not category_slug:
        return JsonResponse({'ok': True, 'results': []})

    products = products.order_by('price', 'name')[:8]
    return JsonResponse(
        {
            'ok': True,
            'component_type': component_type,
            'category_slug': category_slug or '',
            'results': [_bom_product_dict(product) for product in products],
        }
    )


@require_POST
def api_bom_match(request):
    """Принимает список компонентов из схемы, группирует по типу и ищет
    подходящий товар для каждого типа. Возвращает match'и и итоговую цену."""
    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    try:
        matches, grand_total, total_components = _build_bom_matches(payload.get('components', []))
    except ValueError as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)

    return JsonResponse(
        {
            'ok': True,
            'matches': matches,
            'grand_total': grand_total,
            'total_components': total_components,
        }
    )


@require_POST
def api_bom_export_xlsx(request):
    """Build an Excel-compatible BOM workbook for the current schematic."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    try:
        matches, grand_total, total_components = _build_bom_matches(payload.get('components', []))
    except ValueError as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = 'BOM'

    ws.append(['DOLG Bill of Materials'])
    ws.append(['Всего компонентов', total_components, 'Итого, руб.', grand_total])
    ws.append([])

    headers = [
        'Тип',
        'Кол-во',
        'Производитель',
        'Part Number',
        'Наименование',
        'Категория',
        'Корпус',
        'Цена за шт., руб.',
        'Сумма, руб.',
        'Datasheet',
        'Статус',
        'Предупреждения',
    ]
    ws.append(headers)
    header_row = ws.max_row

    for match in matches:
        product = match.get('product')
        if product:
            row = [
                match['label'],
                match['count'],
                product.get('manufacturer') or '',
                product.get('part_number') or '',
                product.get('name') or '',
                product.get('category') or '',
                product.get('package_type') or '',
                product.get('price') or 0,
                match.get('line_total') or 0,
                product.get('datasheet_url') or '',
                product.get('lifecycle_display') or '',
                '; '.join(match.get('warnings') or []),
            ]
        else:
            row = [
                match['label'],
                match['count'],
                '',
                '',
                'нет в каталоге',
                '',
                '',
                '',
                0,
                '',
                '',
                '; '.join(match.get('warnings') or []),
            ]
        ws.append(row)
        datasheet_cell = ws.cell(row=ws.max_row, column=10)
        if datasheet_cell.value:
            datasheet_cell.hyperlink = datasheet_cell.value
            datasheet_cell.style = 'Hyperlink'

    ws.append([])
    ws.append(['', '', '', '', 'ИТОГО', '', '', '', grand_total, '', '', ''])

    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(bold=True)
    ws['C2'].font = Font(bold=True)

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for cell in ws['B']:
        cell.alignment = Alignment(horizontal='center')
    for col_idx in (8, 9):
        for cell in ws[get_column_letter(col_idx)]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    widths = [24, 10, 18, 20, 34, 18, 14, 16, 14, 36, 16, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A{header_row}:L{max(header_row, ws.max_row)}'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="dolg_bom.xlsx"'
    _log_bom_project_event(
        request,
        payload,
        'bom_exported',
        {
            'total_components': total_components,
            'matched_rows': len([item for item in matches if item.get('product')]),
            'grand_total': grand_total,
            'format': 'xlsx',
        },
    )
    return response


@require_POST
def api_bom_add_all(request):
    """Массовое добавление списка товаров в корзину.
    body: {"items": [{"slug": "...", "quantity": 3}, ...]}"""
    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    items = payload.get('items', [])
    if not isinstance(items, list) or not items:
        return JsonResponse({'ok': False, 'error': 'items is empty'}, status=400)
    if len(items) > BOM_MAX_COMPONENTS:
        return JsonResponse(
            {'ok': False, 'error': f'too many items ({len(items)}); limit is {BOM_MAX_COMPONENTS}'},
            status=400,
        )

    cart_filter = _cart_owner_filter(request)
    cart_defaults = _cart_owner_create_defaults(request)
    added = 0
    limited = []
    skipped = []
    added_items = []
    for item in items:
        slug = (item.get('slug') or '').strip()
        try:
            qty = max(1, int(item.get('quantity', 1)))
        except (TypeError, ValueError):
            continue
        if not slug:
            continue
        try:
            product = Product.objects.get(slug=slug)
        except Product.DoesNotExist:
            skipped.append(f'Товар "{slug}" не найден в каталоге.')
            continue
        if product.stock <= 0:
            skipped.append(f'Товар "{product.name}" отсутствует на складе.')
            continue

        cart_item, created = CartItem.objects.get_or_create(
            product=product,
            **cart_filter,
            defaults={'quantity': 0, **cart_defaults},
        )
        available = max(product.stock - cart_item.quantity, 0)
        if available <= 0:
            skipped.append(f'Товар "{product.name}" уже добавлен в пределах складского остатка.')
            continue
        add_qty = min(qty, available)
        if add_qty < qty:
            limited.append(f'Для "{product.name}" добавлено {add_qty} из {qty}: остаток {product.stock} шт.')
        cart_item.quantity += add_qty
        cart_item.save(update_fields=['quantity'])
        added += 1
        added_items.append(
            {
                'slug': product.slug,
                'name': product.name,
                'part_number': product.part_number,
                'quantity': add_qty,
                'unit_price': float(product.price),
                'line_total': round(float(product.price) * add_qty, 2),
            }
        )

    project_context = None
    if added_items:
        project_context = _project_context_from_payload(request, payload, added_items, limited, skipped)
        request.session[PROJECT_CART_SESSION_KEY] = project_context
        request.session.modified = True

    return JsonResponse(
        {
            'ok': True,
            'added': added,
            'limited': limited,
            'skipped': skipped,
            'cart_url': reverse('shop:cart'),
            'project_cart': project_context,
        }
    )


# ---------------------------------------------------------------------------
# PDF-документы: гарантийный лист и сертификат качества (на лету)
# ---------------------------------------------------------------------------

# Кандидаты для cyrillic-шрифта: Windows + Linux (DejaVu из package fonts-dejavu)
# + macOS. Первая найденная пара (regular, bold) выигрывает. Расширяемо.
_CYRILLIC_FONT_CANDIDATES = [
    # Windows
    ('C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/arialbd.ttf'),
    ('C:/Windows/Fonts/times.ttf', 'C:/Windows/Fonts/timesbd.ttf'),
    ('C:/Windows/Fonts/calibri.ttf', 'C:/Windows/Fonts/calibrib.ttf'),
    # Linux (Debian/Ubuntu — fonts-dejavu, fonts-liberation)
    (
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
    ),
    (
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
    ),
    # macOS
    ('/Library/Fonts/Arial.ttf', '/Library/Fonts/Arial Bold.ttf'),
    ('/System/Library/Fonts/Helvetica.ttc', '/System/Library/Fonts/Helvetica.ttc'),
]
_pdf_logger = logging.getLogger(__name__)


def _register_pdf_cyrillic_fonts():
    """Возвращает (regular_name, bold_name). Helvetica-fallback не имеет
    кириллицы — без cyrillic-шрифта получим квадратики в PDF. Поэтому ищем
    в нескольких типичных местах для Win/Linux/macOS, и логируем когда
    fallback на Helvetica реально применился (значит, нужен fonts-пакет в проде)."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Идемпотентность: если уже зарегистрировали в прошлый вызов — переиспользуем.
    if 'DolgCyr' in pdfmetrics.getRegisteredFontNames():
        return 'DolgCyr', 'DolgCyr-Bold'
    for regular_path, bold_path in _CYRILLIC_FONT_CANDIDATES:
        if not Path(regular_path).exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont('DolgCyr', regular_path))
            # Если нет bold-версии — используем regular для bold (визуально хуже,
            # но кириллица сохраняется).
            if Path(bold_path).exists() and bold_path != regular_path:
                pdfmetrics.registerFont(TTFont('DolgCyr-Bold', bold_path))
                return 'DolgCyr', 'DolgCyr-Bold'
            return 'DolgCyr', 'DolgCyr'
        except Exception as exc:
            _pdf_logger.warning('Не удалось зарегистрировать %s: %s', regular_path, exc)
            continue
    _pdf_logger.warning(
        'Cyrillic-шрифт не найден ни в одном из %d кандидатов. PDF будет в Helvetica '
        '(квадратики вместо кириллицы). Установите fonts-dejavu в systeme.',
        len(_CYRILLIC_FONT_CANDIDATES),
    )
    return 'Helvetica', 'Helvetica-Bold'


def _generate_pdf(title, product, sections):
    """Генерирует PDF с обёрткой. `sections` — список (heading, [lines])."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    # Кросс-OS поиск шрифтов с кириллицей — Windows / Linux / macOS. Helvetica-
    # fallback не поддерживает кириллицу, поэтому без cyrillic-шрифта PDF
    # рендерится квадратиками. Берём первый существующий из списка кандидатов.
    font_name, font_bold = _register_pdf_cyrillic_fonts()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle(
        'h1', parent=styles['Title'], fontName=font_bold, fontSize=18, textColor=colors.HexColor('#004080')
    )
    h2 = ParagraphStyle(
        'h2',
        parent=styles['Heading2'],
        fontName=font_bold,
        fontSize=13,
        textColor=colors.HexColor('#002d5a'),
        spaceAfter=6,
    )
    body = ParagraphStyle('body', parent=styles['BodyText'], fontName=font_name, fontSize=10.5, leading=15)
    small = ParagraphStyle(
        'small', parent=styles['BodyText'], fontName=font_name, fontSize=8.5, textColor=colors.grey
    )

    story = [
        Paragraph('ДОЛГ — Магазин электронных компонентов', small),
        Paragraph(title, h1),
        Spacer(1, 0.4 * cm),
        Paragraph(f'<b>Товар:</b> {product.name}', body),
    ]
    if product.part_number:
        story.append(Paragraph(f'<b>Part Number:</b> {product.part_number}', body))
    story.append(Paragraph(f'<b>Артикул:</b> DOLG-{product.id}', body))
    story.append(Paragraph(f'<b>Производитель:</b> {product.get_manufacturer_display()}', body))
    if product.package_type:
        story.append(Paragraph(f'<b>Корпус:</b> {product.package_type}', body))
    story.append(Spacer(1, 0.6 * cm))

    for heading, lines in sections:
        story.append(Paragraph(heading, h2))
        for line in lines:
            story.append(Paragraph(line, body))
        story.append(Spacer(1, 0.5 * cm))

    # Подпись
    story.append(Spacer(1, 1.5 * cm))
    sig_table = Table(
        [['Подпись: ____________________', 'М.П.']],
        colWidths=[9 * cm, 7 * cm],
    )
    sig_table.setStyle(
        TableStyle(
            [
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.append(sig_table)
    story.append(Spacer(1, 0.8 * cm))
    story.append(
        Paragraph(
            f'Документ сгенерирован автоматически. ID заказа: DOLG-{product.id}-{product.slug}',
            small,
        )
    )

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def product_warranty_pdf(request, slug):
    """Генерирует гарантийный талон на товар."""
    product = get_object_or_404(Product, slug=slug)
    sections = [
        (
            'Условия гарантии',
            [
                'Настоящий гарантийный талон подтверждает качество приобретённого товара и право покупателя на бесплатный ремонт или замену в случае выявления производственных дефектов.',
                '<b>Срок гарантии:</b> 12 месяцев с даты приобретения.',
                '<b>Объём гарантии:</b> устранение недостатков, возникших по вине производителя, либо замена товара на аналогичный.',
            ],
        ),
        (
            'Гарантия не распространяется на',
            [
                '• Повреждения, вызванные нарушением правил эксплуатации или монтажа;',
                '• Механические повреждения (сколы, деформация выводов, следы попадания влаги);',
                '• Дефекты, возникшие в результате естественного износа;',
                '• Товары с признаками вскрытия, самостоятельного ремонта или изменения маркировки.',
            ],
        ),
        (
            'Порядок предъявления претензий',
            [
                'Для предъявления претензии обратитесь в службу поддержки ДОЛГ по адресу support@dolg-shop.ru с указанием Part Number и номера заказа. Срок рассмотрения — до 10 рабочих дней.',
            ],
        ),
    ]
    pdf = _generate_pdf(f'Гарантийный талон №{product.id:06d}', product, sections)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="warranty_{product.slug}.pdf"'
    return response


def product_certificate_pdf(request, slug):
    """Генерирует сертификат качества на товар."""
    product = get_object_or_404(Product, slug=slug)
    pn = product.part_number or f'DOLG-{product.id}'
    sections = [
        (
            'Заявление о соответствии',
            [
                f'Настоящим подтверждается, что компонент <b>{pn}</b> производства {product.get_manufacturer_display()} соответствует заявленным техническим характеристикам и допущен к реализации на территории Российской Федерации.',
            ],
        ),
        (
            'Соответствие стандартам',
            [
                '• ГОСТ Р 51317.4.2-2010 — электростатическое разрядное воздействие;',
                '• ГОСТ Р 51318.22-2006 — электромагнитная совместимость;',
                '• IEC 61000-4-x — устойчивость к помехам;',
                '• RoHS 2 (Directive 2011/65/EU) — ограничение использования опасных веществ.',
            ],
        ),
        (
            'Условия применения',
            [
                'Компонент предназначен для использования в составе промышленной электроники, систем РЭБ, радиоприёмных и передающих устройств, средств автоматики. При соблюдении условий эксплуатации (температура, влажность, электрические параметры) обеспечивается заявленный ресурс.',
            ],
        ),
    ]
    pdf = _generate_pdf(f'Сертификат качества №{product.id:06d}/К', product, sections)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="certificate_{product.slug}.pdf"'
    return response


# =============================================================================
# Сравнение товаров — список хранится в session['compare'] (до 4 позиций).
# =============================================================================
MAX_COMPARE = 4


def _compare_slugs(request):
    return request.session.get('compare', [])


@require_POST
def compare_toggle(request, slug):
    """Добавить/убрать товар из списка сравнения. Редиректит обратно на referer."""
    # get_object_or_404 нужен ради 404 на несуществующий slug
    get_object_or_404(Product, slug=slug)
    slugs = _compare_slugs(request)
    if slug in slugs:
        slugs = [s for s in slugs if s != slug]
        added = False
    else:
        if len(slugs) >= MAX_COMPARE:
            from django.contrib import messages

            messages.warning(
                request, f'В сравнении уже {MAX_COMPARE} товара. Удалите один, чтобы добавить новый.'
            )
            return redirect(request.META.get('HTTP_REFERER') or 'shop:index')
        slugs = slugs + [slug]
        added = True
    request.session['compare'] = slugs
    request.session.modified = True

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'added': added, 'count': len(slugs)})
    return redirect(request.META.get('HTTP_REFERER') or 'shop:index')


@require_POST
def compare_clear(request):
    request.session['compare'] = []
    request.session.modified = True
    return redirect('shop:compare')


# --- Анализатор «лучше/хуже» для таблицы сравнения ---
import re as _re

# Метки и ключи параметров, для которых МЕНЬШЕ = ЛУЧШЕ (цена, вес, потребление, задержка).
_LOWER_BETTER_HINTS = (
    'цен',
    'price',
    'руб',
    '₽',
    '$',
    '€',
    'вес',
    'weight',
    'масс',
    'mass',
    'tdp',
    'потребл',
    'consumption',
    'мощност',
    'толщин',
    'thickness',
    'задержк',
    'latency',
    'cas',
    'время',
    'response',
    'esr',
    'rds',
    'утечк',
    'leakage',
)
# Категориальные строки — не сравниваем.
_SKIP_HINTS = (
    'производитель',
    'manufacturer',
    'категори',
    'category',
    'part number',
    'partnumber',
    'корпус',
    'package',
    'lifecycle',
)

_NUM_RE = _re.compile(r'-?\d+(?:[.,]\d+)?')


def _row_compare_kind(label):
    lc = (label or '').lower()
    if any(h in lc for h in _SKIP_HINTS):
        return 'skip'
    if any(h in lc for h in _LOWER_BETTER_HINTS):
        return 'lower'
    return 'higher'


def _extract_number(s):
    """Из строки '8 ГБ', '1.5 ГГц', '120 мАч' → float. Возвращает None если не число."""
    if s is None:
        return None
    m = _NUM_RE.search(str(s))
    if not m:
        return None
    try:
        return float(m.group(0).replace(',', '.'))
    except ValueError:
        return None


def _annotate_row(row):
    """Добавляет row['best_idx'] = индекс лучшей колонки (None если сравнить нельзя)."""
    kind = _row_compare_kind(row['label'])
    if kind == 'skip':
        row['best_idx'] = None
        return row
    nums = [_extract_number(v) for v in row['values']]
    valid = [(i, n) for i, n in enumerate(nums) if n is not None]
    if len(valid) < 2:
        row['best_idx'] = None
        return row
    if len({n for _, n in valid}) == 1:  # ничья
        row['best_idx'] = None
        return row
    best_i, _ = (min if kind == 'lower' else max)(valid, key=lambda p: p[1])
    row['best_idx'] = best_i
    row['compare_kind'] = kind
    return row


def compare_view(request):
    slugs = _compare_slugs(request)
    products_by_slug = {p.slug: p for p in Product.objects.filter(slug__in=slugs).select_related('category')}
    products = [products_by_slug[s] for s in slugs if s in products_by_slug]

    param_keys = []
    seen = set()
    for p in products:
        for k in (p.parameters or {}).keys():
            if k not in seen:
                seen.add(k)
                param_keys.append(k)

    rows = []
    rows.append({'label': 'Производитель', 'values': [p.get_manufacturer_display() for p in products]})
    rows.append({'label': 'Категория', 'values': [p.category.name for p in products]})
    rows.append({'label': 'Цена', 'values': [f'{p.price} ₽' for p in products]})
    rows.append(
        {'label': 'Наличие', 'values': [f'{p.stock} шт.' if p.is_available() else '✕ нет' for p in products]}
    )
    if any(p.part_number for p in products):
        rows.append({'label': 'Part number', 'values': [p.part_number or '—' for p in products]})
    if any(p.lifecycle_status for p in products):
        rows.append({'label': 'Lifecycle', 'values': [p.get_lifecycle_status_display() for p in products]})
    if any(p.package_type for p in products):
        rows.append({'label': 'Корпус', 'values': [p.package_type or '—' for p in products]})
    for key in param_keys:
        label = key.replace('_', ' ').capitalize()
        rows.append(
            {
                'label': label,
                'values': [str((p.parameters or {}).get(key, '—')) for p in products],
            }
        )

    # Аннотируем каждую строку индексом «лучшего» столбца
    rows = [_annotate_row(r) for r in rows]

    # Сводный счётчик побед — для блока «итог по товарам» в шапке
    wins = [0] * len(products)
    for r in rows:
        bi = r.get('best_idx')
        if bi is not None and 0 <= bi < len(wins):
            wins[bi] += 1
    summary = list(zip(products, wins))
    overall_best = max(range(len(products)), key=lambda i: wins[i]) if products else None
    if overall_best is not None and len(set(wins)) == 1:
        overall_best = None  # ничья по всем строкам

    return render(
        request,
        'shop/compare.html',
        {
            'products': products,
            'rows': rows,
            'wins': wins,
            'summary': summary,
            'overall_best': overall_best,
            'max_compare': MAX_COMPARE,
        },
    )


# =============================================================================
# Автодополнение поиска по каталогу (JSON для fetch из шапки).
# =============================================================================
@require_GET
def search_suggest(request):
    """Live-suggest: multi-token + fuzzy fallback (smart_search).

    Юзер набирает «резстор» (опечатка) → fuzzy через rapidfuzz найдёт «резистор».
    Без fuzzy старый icontains-only возвращал [] и suggest исчезал.
    """
    q = (request.GET.get('q') or '').strip()
    if len(q) < 2:
        return JsonResponse({'ok': True, 'results': []})
    base = Product.objects.select_related('category').only(
        'id', 'name', 'slug', 'part_number', 'manufacturer', 'price', 'category__name'
    )
    qs, _tokens = smart_search(base, q)
    qs = qs[:10]
    results = [
        {
            'name': p.name,
            'slug': p.slug,
            'url': f'/product/{p.slug}/',
            'part_number': p.part_number or '',
            'manufacturer': p.get_manufacturer_display(),
            'category': p.category.name,
            'price': float(p.price),
        }
        for p in qs
    ]
    return JsonResponse({'ok': True, 'results': results, 'fuzzy': bool(_tokens) and not bool(results)})


@require_GET
def global_search_suggest(request):
    """Header-search live-suggest. В отличие от каталог-suggest, ищет ВЕЗДЕ:
    товары + категории + статьи энциклопедии + уроки. Лимит по 4 каждой
    группы — чтобы dropdown не вытягивался на половину экрана.
    """
    q = (request.GET.get('q') or '').strip()
    if len(q) < 2:
        return JsonResponse({'ok': True, 'results': []})

    results = []

    # 1) Товары — через smart_search (fuzzy на опечатки)
    base = Product.objects.select_related('category').only(
        'id', 'name', 'slug', 'part_number', 'manufacturer', 'price', 'category__name'
    )
    prod_qs, _ = smart_search(base, q)
    for p in prod_qs[:4]:
        results.append(
            {
                'type': 'product',
                'name': p.name,
                'url': reverse('shop:product_detail', args=[p.slug]),
                'meta': f'{p.get_manufacturer_display()} · {p.category.name} · {int(p.price)} ₽',
                'icon': '🛒',
            }
        )

    # 2) Категории — простой icontains
    for cat in Category.objects.filter(Q(name__icontains=q) | Q(slug__icontains=q))[:3]:
        results.append(
            {
                'type': 'category',
                'name': cat.name,
                'url': reverse('shop:category', args=[cat.slug]),
                'meta': 'Категория каталога',
                'icon': '📂',
            }
        )

    # 3) Статьи энциклопедии (если приложение доступно)
    try:
        from knowledge.models import Article

        articles = (
            Article.objects.select_related('category')
            .filter(is_published=True)
            .filter(Q(title__icontains=q) | Q(summary__icontains=q) | Q(category__name__icontains=q))[:3]
        )
        for a in articles:
            results.append(
                {
                    'type': 'article',
                    'name': a.title,
                    'url': reverse('knowledge:article', args=[a.slug]),
                    'meta': f'Энциклопедия · {a.category.name}' if a.category_id else 'Энциклопедия',
                    'icon': '📚',
                }
            )
    except Exception:
        pass

    # 4) Официальные источники и документация
    for source in _legal_source_results(q, limit=3):
        results.append(
            {
                'type': 'legal_source',
                'name': source['title'],
                'url': source['article_url'],
                'meta': f'Источник · {source.get("topic", "")}',
                'icon': '🔎',
            }
        )

    # 5) Уроки (если есть)
    try:
        from knowledge.models import LearningLesson

        lessons = LearningLesson.objects.filter(
            Q(title__icontains=q) | Q(summary__icontains=q)
        ).select_related('track')[:2]
        for lesson in lessons:
            results.append(
                {
                    'type': 'lesson',
                    'name': lesson.title,
                    'url': reverse('knowledge:learning_lesson', args=[lesson.slug]),
                    'meta': f'Обучение · {lesson.track.title}'
                    if getattr(lesson, 'track_id', None)
                    else 'Обучение',
                    'icon': '🎓',
                }
            )
    except Exception:
        pass

    return JsonResponse({'ok': True, 'results': results})
