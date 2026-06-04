import json
import tempfile
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from PIL import Image, ImageDraw

from .models import Category, Product


class CategoryModelTest(TestCase):
    def setUp(self):
        self.category = Category.objects.create(
            name='Smartphones', description='Mobile phones and accessories'
        )

    def test_category_creation(self):
        self.assertEqual(self.category.name, 'Smartphones')
        self.assertTrue(self.category.slug)

    def test_category_str(self):
        self.assertEqual(str(self.category), 'Smartphones')


class ProductModelTest(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name='Electronics')
        self.product = Product.objects.create(
            name='iPhone 14',
            category=self.category,
            description='Latest Apple smartphone',
            price=999.99,
            stock=10,
            manufacturer='apple',
        )

    def test_product_creation(self):
        self.assertEqual(self.product.name, 'iPhone 14')
        self.assertEqual(self.product.price, Decimal('999.99'))
        self.assertTrue(self.product.is_available())


class ProductCardHelperTests(TestCase):
    def test_parameter_preview_hides_service_metadata(self):
        from shop.card_helpers import group_parameters, parameter_preview

        params = {
            'current': '5.7 А',
            'max_voltage': '30 В',
            'type': 'N-MOSFET',
            'catalog_quality': {'normalizer': 'reb_catalog_quality_v1'},
            'image_verified_from': 'supplier cache',
        }

        preview = parameter_preview(params, limit=5)
        keys = {item['key'] for item in preview}
        grouped_text = json.dumps(group_parameters(params), ensure_ascii=False)

        self.assertEqual(keys, {'current', 'max_voltage', 'type'})
        self.assertNotIn('catalog_quality', grouped_text)
        self.assertNotIn('image_verified_from', grouped_text)
        self.assertEqual(next(item for item in preview if item['key'] == 'max_voltage')['label'], 'Umax')

    def test_parameter_preview_marks_long_values_as_wide(self):
        from shop.card_helpers import parameter_preview

        preview = parameter_preview({'type': 'Кабель HDMI 2.1 (Ultra High Speed)'}, limit=1)

        self.assertTrue(preview[0]['wide'])

    def test_parameter_preview_does_not_duplicate_repeated_order_keys(self):
        from shop.card_helpers import parameter_preview

        preview = parameter_preview(
            {
                'type': 'digital oscilloscope',
                'interface': 'USB/LAN',
                'channels': '4',
                'sample_rate': '1 Гвыб/с',
            },
            limit=8,
        )

        keys = [item['key'] for item in preview]
        self.assertEqual(keys.count('interface'), 1)


class EnrichProductParametersTests(TestCase):
    def test_command_fills_module_specs_without_removing_image_metadata(self):
        category = Category.objects.create(name='Modules', slug='modules')
        product = Product.objects.create(
            name='Arduino Uno R3',
            slug='arduino-uno-r3',
            category=category,
            description='Development board.',
            price=1000,
            stock=10,
            manufacturer='other',
            parameters={'image_source': 'verified', 'image_source_policy': 'official'},
        )

        call_command('enrich_product_parameters', stdout=StringIO())
        product.refresh_from_db()

        self.assertEqual(product.parameters['mcu'], 'ATmega328P')
        self.assertEqual(product.parameters['logic_level'], '5 В')
        self.assertEqual(product.parameters['image_source'], 'verified')

    def test_command_repairs_known_seed_ssd_artifact(self):
        category = Category.objects.create(name='SSD', slug='ssd')
        product = Product.objects.create(
            name='T1 OK',
            slug='t1-ok',
            category=category,
            description='Test',
            price=1,
            stock=1,
            manufacturer='other',
            parameters={'image_source': 'generated'},
        )

        call_command('enrich_product_parameters', stdout=StringIO())
        product.refresh_from_db()

        self.assertEqual(product.name, 'Samsung Portable SSD T7 1TB')
        self.assertEqual(product.manufacturer, 'samsung')
        self.assertEqual(product.parameters['interface'], 'USB 3.2 Gen 2 (USB-C)')
        self.assertEqual(product.parameters['image_source'], 'generated')

    def test_command_repairs_consumable_seed_text_and_package(self):
        category = Category.objects.create(name='Consumables', slug='consumables')
        product = Product.objects.create(
            name='BREADBOARD 2x830',
            slug='breadboard-2x830',
            category=category,
            description='Seed breadboard item.',
            price=780,
            stock=500,
            manufacturer='other',
            package_type='Consumable',
            parameters={'type': 'breadboard'},
        )

        call_command('enrich_product_parameters', stdout=StringIO())
        product.refresh_from_db()

        self.assertEqual(product.name, 'Набор макетных плат 2×830')
        self.assertEqual(product.package_type, 'Макетирование')
        self.assertEqual(product.parameters['type'], 'Набор макетных плат')


class DatasheetIntelligenceTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name='Resistors', slug='resistors')
        self.product = Product.objects.create(
            name='Vishay CRCW0805 10k',
            category=self.category,
            description='Thick film chip resistor for voltage dividers and pull-ups.',
            price=3.5,
            stock=100,
            manufacturer='vishay',
            part_number='CRCW080510K0FKEA',
            package_type='0805',
            datasheet_url='https://example.com/crcw0805.pdf',
            parameters={
                'resistance': '10 кОм',
                'tolerance': '1%',
                'power': '0.125 Вт',
                'voltage': '150 В',
                'mounting': 'SMD',
            },
        )

    def test_product_metadata_builds_datasheet_record_fields(self):
        from shop.services.datasheet_intelligence import build_product_datasheet_record

        record = build_product_datasheet_record(self.product)
        fields = record['fields']

        self.assertGreaterEqual(record['confidence'], 0.5)
        self.assertIn('package', record['metadata_inferred_fields'])
        self.assertTrue(fields['absolute_maximum_ratings'])
        self.assertTrue(fields['recommended_operating_conditions'])

    def test_enrich_datasheets_command_fills_missing_records(self):
        out = StringIO()

        call_command('enrich_datasheets', '--all', '--missing-only', '--json', stdout=out)
        self.product.refresh_from_db()

        record = self.product.parameters.get('datasheet_extracted')
        self.assertIsInstance(record, dict)
        self.assertIn('fields', record)
        self.assertTrue(record['fields']['package'])

    def test_product_detail_shows_datasheet_intelligence_summary(self):
        from shop.services.datasheet_intelligence import build_product_datasheet_record

        params = dict(self.product.parameters)
        params['datasheet_extracted'] = build_product_datasheet_record(self.product)
        self.product.parameters = params
        self.product.save(update_fields=['parameters'])

        response = self.client.get(reverse('shop:product_detail', args=[self.product.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Datasheet Intelligence')
        self.assertContains(response, 'Предельные режимы')
        self.assertContains(response, 'Рабочие условия')
        self.assertNotContains(response, 'datasheet_extracted')


class RebCatalogQualityTests(TestCase):
    def setUp(self):
        self.connectors = Category.objects.create(name='Connectors', slug='connectors')
        self.diodes = Category.objects.create(name='Diodes', slug='diodes')

    def test_normalizer_fills_connector_part_number_mounting_and_ratings(self):
        from shop.services.reb_catalog_quality import normalize_reb_product

        product = Product.objects.create(
            name='JST XH 2P',
            slug='jst-xh-2p',
            category=self.connectors,
            description='JST XH 2-pin board connector.',
            price=10,
            stock=100,
            manufacturer='other',
            package_type='SMD',
            parameters={},
        )

        result = normalize_reb_product(product)

        self.assertTrue(result.changed)
        self.assertEqual(product.part_number, 'JST-XH-2P')
        self.assertEqual(product.parameters['mounting'], 'THT')
        self.assertEqual(product.parameters['current'], '3 А')
        self.assertIn('jst-mfg.com', product.datasheet_url)

    def test_normalizer_fills_diode_engineering_fields(self):
        from shop.services.reb_catalog_quality import normalize_reb_product

        product = Product.objects.create(
            name='onsemi 1N4007',
            slug='1n4007',
            category=self.diodes,
            description='Rectifier diode.',
            price=5,
            stock=100,
            manufacturer='onsemi',
            part_number='1N4007',
            package_type='DO-35',
            parameters={},
        )

        normalize_reb_product(product)

        self.assertEqual(product.package_type, 'THT DO-41')
        self.assertEqual(product.parameters['mounting'], 'THT')
        self.assertEqual(product.parameters['max_voltage'], '1000 В')
        self.assertEqual(product.parameters['current'], '1 А')
        self.assertIn('onsemi.com', product.datasheet_url)

    def test_dry_run_command_does_not_save_changes(self):
        product = Product.objects.create(
            name='DB9 female',
            slug='db9-female-tht',
            category=self.connectors,
            description='D-Sub connector.',
            price=45,
            stock=20,
            manufacturer='other',
            package_type='THT',
            parameters={},
        )

        call_command('normalize_reb_catalog', '--dry-run', verbosity=0)
        product.refresh_from_db()

        self.assertEqual(product.part_number, '')
        self.assertEqual(product.parameters, {})


class ShopViewsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.category = Category.objects.create(name='Laptops')
        self.product = Product.objects.create(
            name='MacBook Pro',
            category=self.category,
            description='Professional laptop',
            price=1999.99,
            stock=5,
            manufacturer='apple',
        )

    def test_index_view(self):
        response = self.client.get(reverse('shop:index'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'MacBook Pro')

    def test_product_detail_view(self):
        response = self.client.get(reverse('shop:product_detail', args=[self.product.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'MacBook Pro')
        self.assertContains(response, 'Инженерное применение')
        self.assertContains(response, 'Схемы')


class CompareAnalyzerTests(TestCase):
    """Анализатор «лучше/хуже» для таблицы сравнения товаров."""

    def test_lower_better_recognizes_price_keywords(self):
        from shop.views import _row_compare_kind

        # Цена и веса — «меньше = лучше»
        self.assertEqual(_row_compare_kind('Цена'), 'lower')
        self.assertEqual(_row_compare_kind('Цена, ₽'), 'lower')
        self.assertEqual(_row_compare_kind('Стоимость, руб'), 'lower')
        self.assertEqual(_row_compare_kind('Price'), 'lower')
        self.assertEqual(_row_compare_kind('Масса, г'), 'lower')
        self.assertEqual(_row_compare_kind('TDP, Вт'), 'lower')
        self.assertEqual(_row_compare_kind('Потребление, мА'), 'lower')

    def test_higher_better_default(self):
        from shop.views import _row_compare_kind

        # Тактовая, объём — «больше = лучше»
        self.assertEqual(_row_compare_kind('Тактовая частота, ГГц'), 'higher')
        self.assertEqual(_row_compare_kind('RAM, ГБ'), 'higher')

    def test_skip_categorical(self):
        from shop.views import _row_compare_kind

        # Производитель и категория — не сравниваются
        self.assertEqual(_row_compare_kind('Производитель'), 'skip')
        self.assertEqual(_row_compare_kind('Корпус'), 'skip')
        self.assertEqual(_row_compare_kind('Lifecycle'), 'skip')


class CatalogFilterTests(TestCase):
    """_apply_filters — общая фильтрация каталога по GET-параметрам."""

    def setUp(self):
        cat = Category.objects.create(name='Резисторы', slug='resistors')
        # slug передаём явно: Product.save() делает slugify без allow_unicode,
        # для кириллицы это даст пустой slug и {% url %} падает в шаблоне.
        Product.objects.create(
            name='Резистор 1к',
            slug='resistor-1k',
            category=cat,
            description='metal-film',
            price=5,
            stock=100,
            manufacturer='vishay',
            part_number='MF-1K',
            lifecycle_status='active',
            package_type='SMD-0805',
            parameters={'resistance': '1k', 'power': '0.25 W', 'voltage': '50 V', 'tolerance': '1%'},
        )
        Product.objects.create(
            name='Резистор 10к',
            slug='resistor-10k',
            category=cat,
            description='precision',
            price=8,
            stock=100,
            manufacturer='yageo',
            part_number='RC-10K',
            lifecycle_status='nrnd',
            package_type='SMD-1206',
            parameters={'resistance': '10k', 'power': '0.125 W', 'voltage': '25 V', 'tolerance': '5%'},
        )
        Product.objects.create(
            name='Резистор устаревший',
            slug='resistor-obsolete',
            category=cat,
            description='legacy',
            price=2,
            stock=0,
            manufacturer='vishay',
            part_number='OLD-R',
            lifecycle_status='obsolete',
            package_type='THT',
        )
        self.cat = cat

    def test_q_filter_matches_part_number(self):
        resp = self.client.get(reverse('shop:index') + '?q=MF-1K')
        self.assertContains(resp, 'Резистор 1к')
        self.assertNotContains(resp, 'Резистор 10к')

    def test_q_filter_matches_description(self):
        resp = self.client.get(reverse('shop:index') + '?q=precision')
        self.assertContains(resp, 'Резистор 10к')
        self.assertNotContains(resp, 'Резистор 1к')

    def test_manufacturer_filter(self):
        resp = self.client.get(reverse('shop:index') + '?manufacturer=yageo')
        self.assertContains(resp, 'Резистор 10к')
        self.assertNotContains(resp, 'Резистор 1к')

    def test_lifecycle_filter(self):
        resp = self.client.get(reverse('shop:index') + '?lifecycle=obsolete')
        self.assertContains(resp, 'Резистор устаревший')
        self.assertNotContains(resp, 'Резистор 1к')

    def test_package_filter_substring(self):
        # «SMD» должен сматчить и SMD-0805, и SMD-1206
        resp = self.client.get(reverse('shop:index') + '?package=SMD')
        self.assertContains(resp, 'Резистор 1к')
        self.assertContains(resp, 'Резистор 10к')
        self.assertNotContains(resp, 'Резистор устаревший')

    def test_combined_filters(self):
        resp = self.client.get(reverse('shop:index') + '?manufacturer=vishay&lifecycle=active')
        self.assertContains(resp, 'Резистор 1к')
        self.assertNotContains(resp, 'Резистор 10к')
        self.assertNotContains(resp, 'Резистор устаревший')

    def test_engineering_nominal_filter(self):
        resp = self.client.get(reverse('shop:index') + '?nominal=10k')
        self.assertContains(resp, 'Резистор 10к')
        self.assertNotContains(resp, 'Резистор 1к')

    def test_engineering_power_voltage_tolerance_filters(self):
        resp = self.client.get(reverse('shop:index') + '?power=0.25&voltage=50&tolerance=1%25')
        self.assertContains(resp, 'Резистор 1к')
        self.assertNotContains(resp, 'Резистор 10к')

    def test_extended_chip_filters_have_backend_support(self):
        from shop.card_helpers import CARD_PARAM_FILTER_MAP, chip_filter_name, chip_value_valid
        from shop.services.catalog_filters import CATALOG_FILTERS

        self.assertEqual(chip_filter_name('current'), 'current')
        self.assertEqual(chip_filter_name('screen_size'), 'display')
        self.assertEqual(chip_filter_name('package'), 'package')
        self.assertEqual(chip_filter_name('material'), 'material')
        self.assertEqual(chip_filter_name('contact_material'), 'material')
        self.assertEqual(chip_filter_name('temperature_range'), 'temperature_range')
        self.assertEqual(chip_filter_name('length'), 'size')
        self.assertEqual(chip_filter_name('power_rails'), 'size')
        self.assertEqual(chip_filter_name('gauge'), 'wire')
        self.assertEqual(chip_filter_name('configuration'), 'configuration')
        self.assertEqual(chip_filter_name('latency'), 'q')
        self.assertTrue(chip_value_valid({}, 'latency', 'CL34'))
        self.assertFalse(set(CARD_PARAM_FILTER_MAP.values()) - set(CATALOG_FILTERS))

    def test_in_stock_filter_hides_zero_stock(self):
        resp = self.client.get(reverse('shop:index') + '?in_stock=1')
        self.assertContains(resp, 'Резистор 1к')
        self.assertNotContains(resp, 'Резистор устаревший')


class SearchSuggestTests(TestCase):
    """search_suggest — JSON-автодополнение поиска."""

    def setUp(self):
        cat = Category.objects.create(name='ICs', slug='ics')
        Product.objects.create(
            name='STM32F103C8T6',
            category=cat,
            description='ARM Cortex-M3',
            price=120,
            stock=50,
            manufacturer='st',
            part_number='STM32F103C8T6',
        )

    def test_returns_empty_for_short_query(self):
        resp = self.client.get(reverse('shop:search_suggest') + '?q=s')
        data = resp.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['results'], [])

    def test_finds_by_part_number(self):
        resp = self.client.get(reverse('shop:search_suggest') + '?q=STM32')
        data = resp.json()
        self.assertEqual(len(data['results']), 1)
        self.assertEqual(data['results'][0]['part_number'], 'STM32F103C8T6')

    def test_response_shape(self):
        resp = self.client.get(reverse('shop:search_suggest') + '?q=ARM')
        item = resp.json()['results'][0]
        for key in ('name', 'slug', 'url', 'part_number', 'manufacturer', 'category', 'price'):
            self.assertIn(key, item)


class GlobalSearchAndDemoRouteTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name='Микросхемы', slug='ics')
        self.product = Product.objects.create(
            name='NE555 timer',
            slug='ne555-timer',
            category=self.category,
            description='Таймер для генераторов и задержек',
            price=25,
            stock=100,
            manufacturer='ti',
            part_number='NE555DR',
        )

        from knowledge.models import Article, KnowledgeCategory, LearningLesson, LearningTask, LearningTrack

        self.kb_category = KnowledgeCategory.objects.create(
            name='Практика',
            slug='practice',
            topic='practice',
        )
        self.article = Article.objects.create(
            category=self.kb_category,
            title='Таймер NE555 в генераторе',
            slug='ne555-generator',
            summary='Как применять NE555 в простой схеме генератора.',
            body='<p>NE555 подходит для генератора импульсов.</p>',
            is_published=True,
        )
        self.learning_track = LearningTrack.objects.create(
            title='Базовая схемотехника',
            slug='basic-electronics',
            summary='Практические задания по расчетам и схемам.',
            level='basic',
        )
        self.learning_lesson = LearningLesson.objects.create(
            track=self.learning_track,
            title='Закон Ома и делитель',
            slug='ohm-divider-learning',
            summary='Закон Ома, делитель и RC как практические задачи.',
            theory='<p>Соберите делитель и измерьте выход.</p>',
            formula='I = U / R',
        )
        LearningTask.objects.create(
            lesson=self.learning_lesson,
            task_type='math_numeric',
            title='Задача на делитель',
            prompt='Рассчитайте выход делителя напряжения.',
            rubric={'expected_value': 2.88, 'unit': 'В', 'tolerance_abs': 0.05},
        )
        self.lab_lesson = LearningLesson.objects.create(
            track=self.learning_track,
            title='Генератор NE555',
            slug='ne555-learning',
            summary='Расчет частоты NE555 и проверка duty cycle.',
            theory='<p>NE555 работает как астабильный генератор.</p>',
            formula='f = 1.44 / ((R1 + 2R2)C)',
            order=20,
        )
        LearningTask.objects.create(
            lesson=self.lab_lesson,
            task_type='simulation_measure',
            title='Измерить частоту NE555',
            prompt='Найдите частоту и duty cycle генератора NE555.',
            rubric={'expected_value': 100, 'unit': 'Гц', 'tolerance_abs': 10},
        )

        from Dolg_APP.models import SchematicProject

        user = get_user_model().objects.create_user(username='demo-owner', password='pass')
        self.project = SchematicProject.objects.create(
            user=user,
            name='NE555 demo astable',
            description='Демо-схема мультивибратора на NE555',
            is_demo=True,
            scheme_data={'components': [], 'connections': []},
        )

    def test_global_search_finds_products_articles_projects_and_tools(self):
        response = self.client.get(reverse('shop:global_search') + '?q=NE555')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'NE555 timer')
        self.assertContains(response, 'Таймер NE555 в генераторе')
        self.assertContains(response, 'NE555 demo astable')

    def test_global_search_finds_tool_by_keyword(self):
        response = self.client.get(reverse('shop:global_search') + '?q=BOM')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Корзина и BOM')

    def test_global_search_finds_learning_lessons_and_tasks(self):
        response = self.client.get(reverse('shop:global_search') + '?q=закон ома')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Закон Ома и делитель')
        self.assertContains(response, 'Обучение')

    def test_global_search_finds_engineering_lab_topics(self):
        response = self.client.get(reverse('shop:global_search') + '?q=NE555')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Инженерная лаборатория')
        self.assertContains(response, 'Генератор NE555')

    def test_global_search_finds_review_and_import_topics(self):
        response = self.client.get(reverse('shop:global_search') + '?q=LTspice')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'CAD Import to Review')

        response = self.client.get(reverse('shop:global_search') + '?q=derating')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Engineering Review')

    def test_global_search_finds_expert_system_topics(self):
        for query in ('rule-engine', 'jsonschema', 'pint', 'z3', 'fuzzy'):
            response = self.client.get(reverse('shop:global_search') + f'?q={query}')
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, 'Expert System Review')

    def test_global_search_finds_pro_simulation_analytics_topics(self):
        for query in ('FFT', 'Bode', 'Monte Carlo', 'SciPy', 'THD', 'sweep'):
            response = self.client.get(reverse('shop:global_search') + f'?q={query}')
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, 'Pro-аналитика симуляции')

    def test_global_search_finds_legal_sources(self):
        response = self.client.get(reverse('shop:global_search') + '?q=ngspice')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Источники и документация')
        self.assertContains(response, 'ngspice Documentation')

    def test_lookup_suggests_legal_sources(self):
        response = self.client.get(reverse('shop:global_search_suggest') + '?q=PyTorch')

        self.assertEqual(response.status_code, 200)
        results = response.json()['results']
        self.assertTrue(any(item['type'] == 'legal_source' for item in results))

    def test_demo_route_renders_integrated_path(self):
        response = self.client.get(reverse('shop:demo_route'))
        self.assertEqual(response.status_code, 200)
        for word in ('Каталог', 'Энциклопедия', 'CAD', 'Симуляция', 'BOM', 'Заказ'):
            self.assertContains(response, word)


class DemoReadyCommandScientificStackTests(TestCase):
    def test_scientific_stack_smoke_checks_dependencies_and_services(self):
        from shop.management.commands.check_demo_ready import Command

        report = {
            'ok': True,
            'errors': [],
            'warnings': [],
            'counts': {},
            'urls': [],
            'scientific_stack': {},
            'expert_stack': {},
            'project_session': {},
            'legal_sources_stack': {},
        }

        Command()._check_scientific_stack(report)
        Command()._check_project_session_stack(report)
        Command()._check_expert_system_stack(report)

        self.assertEqual(report['errors'], [])
        stack = report['scientific_stack']
        for package_name in ('numpy', 'scipy', 'matplotlib', 'pandas', 'python-engineering'):
            self.assertTrue(stack[package_name]['ok'], package_name)
            self.assertTrue(stack[package_name]['version'])

        for check_name in (
            'fft_svg',
            'bode_svg',
            'monte_carlo_svg',
            'signal_quality_svg',
            'parameter_sweep_svg',
            'dc_fallback',
        ):
            self.assertTrue(stack['service_checks'][check_name], check_name)

        session = report['project_session']
        for check_name in (
            'project_event_model',
            'simulation_async_fields',
            'postprocess_measurements',
            'csv_export',
            'validity_guard',
        ):
            self.assertTrue(session['service_checks'][check_name], check_name)

        expert = report['expert_stack']
        for package_name in ('jsonschema', 'rule-engine', 'pint', 'lark', 'z3-solver', 'scikit-fuzzy'):
            self.assertTrue(expert[package_name]['ok'], package_name)
            self.assertTrue(expert[package_name]['version'])
        for check_name in (
            'rule_pack',
            'rule_engine_finding',
            'pint_unit_parse',
            'lark_import_preview',
            'cad_import_preview_details',
            'learning_by_review',
            'z3_voltage_divider',
            'fuzzy_risk',
        ):
            self.assertTrue(expert['service_checks'][check_name], check_name)

    def test_legal_sources_stack_checks_retrieval_rules_search_and_learning(self):
        from shop.management.commands.check_demo_ready import Command

        call_command('seed_legal_sources', verbosity=0)
        report = {
            'ok': True,
            'errors': [],
            'warnings': [],
            'legal_sources_stack': {},
        }

        Command()._check_legal_sources_stack(report)

        self.assertEqual(report['errors'], [])
        stack = report['legal_sources_stack']
        self.assertTrue(stack['service_checks']['source_retrieval'])
        self.assertTrue(stack['service_checks']['rule_bibliography'])
        self.assertTrue(stack['service_checks']['search_smoke'])
        self.assertGreaterEqual(stack['learning_tasks_with_sources'], 8)


class DataIntegrityLegalSourcesTests(TestCase):
    def test_unknown_learning_source_id_is_reported(self):
        from knowledge.models import LearningLesson, LearningTask, LearningTrack
        from shop.management.commands.check_data_integrity import Command

        track = LearningTrack.objects.create(
            title='Broken source track',
            slug='broken-source-track',
            summary='Demo',
        )
        lesson = LearningLesson.objects.create(
            track=track,
            title='Broken source lesson',
            slug='broken-source-lesson',
            summary='Demo',
            theory='Demo',
        )
        LearningTask.objects.create(
            lesson=lesson,
            task_type='math_numeric',
            title='Broken source task',
            prompt='Demo',
            rubric={'expected_value': 1, 'source_ids': ['missing_source']},
        )
        report = {
            'ok': True,
            'counts': {},
            'errors': [],
            'warnings': [],
            'catalog': {},
            'knowledge': {},
            'schematics': {},
            'legal_sources': {},
        }

        Command()._check_legal_sources(report)

        self.assertTrue(report['legal_sources']['task_source_errors'])
        self.assertTrue(
            any('learning tasks reference unknown legal sources' in item for item in report['errors'])
        )


class ProductImagePolicyTests(TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.override = override_settings(MEDIA_ROOT=self.tmp.name)
        self.override.enable()
        self.category = Category.objects.create(name='ICs', slug='ics')
        self.product = Product.objects.create(
            name='ST LM358DT',
            slug='st-lm358dt',
            category=self.category,
            description='Dual operational amplifier',
            price=32,
            stock=10,
            manufacturer='st',
            part_number='LM358DT',
            package_type='SMD SOIC-8',
        )

    def tearDown(self):
        self.override.disable()
        self.tmp.cleanup()

    def test_generator_assigns_local_png_and_source_policy(self):
        from shop.services.product_images import apply_generated_product_image

        changed = apply_generated_product_image(self.product, force=True)
        self.product.refresh_from_db()

        self.assertTrue(changed)
        self.assertTrue(self.product.image.name.startswith('products/generated/'))
        self.assertTrue(self.product.image.name.endswith('.png'))
        self.assertEqual(self.product.parameters['image_source_policy'], 'no-wikimedia')
        self.assertIn('generated-product-art', self.product.parameters['image_source_url'])
        self.assertTrue(self.product.image.path.endswith('.png'))

    def test_management_command_replaces_commons_image(self):
        self.product.image.name = 'products/commons/st-lm358dt.jpg'
        self.product.parameters = {'image_source_url': 'https://commons.wikimedia.org/wiki/File:Khao_soi.jpg'}
        self.product.save(update_fields=['image', 'parameters'])

        call_command('apply_curated_product_photos', '--force', verbosity=0)
        self.product.refresh_from_db()

        self.assertTrue(self.product.image.name.startswith('products/generated/'))
        self.assertNotIn('wikimedia', self.product.parameters['image_source_url'].lower())

    def test_management_command_preserves_exact_local_asset(self):
        # Raster asset имеет высший приоритет: если для slug есть точный PNG,
        # policy не должна оставлять generated-заглушку.
        local_asset = Path(self.tmp.name) / 'products' / 'st-lm358dt.png'
        local_asset.parent.mkdir(parents=True, exist_ok=True)
        # 1x1 PNG (simplest valid file). Pillow читает и Image.open()-нет.
        local_asset.write_bytes(
            b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01'
            b'\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\rIDATx\x9cc\xf8'
            b'\xcf\xc0\x00\x00\x00\x03\x00\x01\xc7\xd7\xa3w\x00\x00\x00\x00IEND\xaeB`\x82'
        )
        # Set image to generated; command should re-choose to local PNG.
        self.product.image.name = 'products/generated/st-lm358dt.png'
        self.product.save(update_fields=['image'])

        call_command('apply_curated_product_photos', verbosity=0)
        self.product.refresh_from_db()

        self.assertEqual(self.product.image.name, 'products/st-lm358dt.png')
        self.assertEqual(self.product.parameters['image_source_url'], 'local://dolg/product-asset')

    def test_management_command_uses_generated_ugo_when_only_svg_asset_exists(self):
        from shop.services.media_quality import audit_product_image

        local_asset = Path(self.tmp.name) / 'products' / 'st-lm358dt.svg'
        local_asset.parent.mkdir(parents=True, exist_ok=True)
        local_asset.write_text(
            '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 120 80">'
            '<rect width="120" height="80" fill="#081224"/>'
            '<rect x="32" y="20" width="56" height="40" rx="6" fill="#111827" stroke="#00d4ff"/>'
            '<text x="60" y="46" text-anchor="middle" fill="#e8f6ff" font-size="14">IC</text>'
            '</svg>',
            encoding='utf-8',
        )
        self.product.image.name = 'products/generated/st-lm358dt.png'
        self.product.save(update_fields=['image'])

        call_command('apply_curated_product_photos', verbosity=0)
        self.product.refresh_from_db()

        self.assertEqual(self.product.image.name, 'products/generated/st-lm358dt.png')
        self.assertEqual(
            self.product.parameters['image_source'],
            'manufacturer image pending; generated technical placeholder',
        )
        report = audit_product_image(self.product)
        self.assertTrue(report['ok'], report)
        self.assertEqual(report['source_type'], 'generated')
        self.assertEqual(report['metrics']['format'], 'png')
        self.assertEqual(report['warnings'], [])

    def test_verified_photo_command_applies_real_photo_and_skips_blocked_slug(self):
        from shop.services.product_images import VERIFIED_IMAGE_SOURCE

        cpu_cat = Category.objects.create(name='CPU', slug='cpu')
        cpu = Product.objects.create(
            name='AMD Ryzen 5 7600X',
            slug='amd-ryzen-5-7600x',
            category=cpu_cat,
            description='Desktop CPU',
            price=18000,
            stock=3,
            manufacturer='amd',
        )
        source_dir = Path(self.tmp.name) / 'products' / 'commons'
        source_dir.mkdir(parents=True, exist_ok=True)
        photo = Image.new('RGB', (640, 480), (80, 120, 160))
        draw = ImageDraw.Draw(photo)
        for x in range(0, 640, 32):
            draw.line((x, 0, 640 - x // 2, 479), fill=(220, 230, 240), width=3)
        draw.rectangle((190, 130, 450, 350), fill=(30, 40, 55), outline=(230, 230, 230), width=8)
        photo.save(source_dir / 'amd-ryzen-5-7600x.jpg', format='JPEG')
        Image.new('RGB', (640, 480), (120, 80, 60)).save(source_dir / 'st-lm358dt.jpg', format='JPEG')

        call_command('apply_verified_product_photos', verbosity=0)
        cpu.refresh_from_db()
        self.product.refresh_from_db()

        self.assertEqual(cpu.image.name, 'products/verified/amd-ryzen-5-7600x.jpg')
        self.assertEqual(cpu.parameters['image_source_url'], VERIFIED_IMAGE_SOURCE)
        self.assertEqual(cpu.parameters['image_source'], 'verified real product photo')
        self.assertNotEqual(self.product.image.name, 'products/verified/st-lm358dt.jpg')

    def test_legacy_commons_import_is_disabled(self):
        with self.assertRaises(CommandError):
            call_command('import_commons_product_photos', verbosity=0)

    def test_template_filter_blocks_legacy_external_sources(self):
        # После split-1 (2026-05-19) функция переехала из templatetags-модуля
        # в shop.card_helpers; templatetags теперь тонкий register-layer.
        from shop.card_helpers import is_stock_image

        class FakeImage:
            name = 'products/commons/st-lm358dt.jpg'

        self.assertTrue(is_stock_image(FakeImage()))

    def test_policy_detects_forbidden_path(self):
        from shop.services.product_images import is_forbidden_image_path

        self.assertTrue(is_forbidden_image_path('products/commons/foo.jpg'))
        self.assertTrue(is_forbidden_image_path('https://upload.wikimedia.org/foo.jpg'))
        self.assertFalse(is_forbidden_image_path('products/generated/foo.png'))

    def test_generated_product_image_passes_media_quality_gate(self):
        from shop.services.media_quality import audit_product_image
        from shop.services.product_images import apply_generated_product_image

        apply_generated_product_image(self.product, force=True)
        self.product.refresh_from_db()

        report = audit_product_image(self.product)
        self.assertTrue(report['ok'], report)
        self.assertEqual(report['source_type'], 'generated')
        self.assertGreaterEqual(report['metrics']['width'], 320)
        self.assertGreaterEqual(report['metrics']['height'], 220)
        self.assertTrue(report['hashes']['available'])
        self.assertTrue(report['hashes']['perceptual_hash'])

    def test_media_quality_gate_rejects_tiny_local_asset(self):
        from shop.services.media_quality import audit_product_image

        local_asset = Path(self.tmp.name) / 'products' / 'st-lm358dt.png'
        local_asset.parent.mkdir(parents=True, exist_ok=True)
        Image.new('RGB', (1, 1), (255, 255, 255)).save(local_asset, format='PNG')
        self.product.image.name = 'products/st-lm358dt.png'
        self.product.save(update_fields=['image'])

        report = audit_product_image(self.product)
        self.assertFalse(report['ok'])
        self.assertEqual(report['source_type'], 'local_asset')
        self.assertIn('image_too_small', report['errors'])
        self.assertIn('image_near_blank', report['errors'])


class ComponentSearchTests(TestCase):
    """api_component_search — подбор товара каталога для выбранного элемента схемы."""

    def setUp(self):
        self.cat_r = Category.objects.create(name='Резисторы', slug='resistors')
        self.cat_t = Category.objects.create(name='Транзисторы', slug='transistors')
        Product.objects.create(
            name='Резистор 1к',
            slug='res-1k',
            category=self.cat_r,
            description='1k',
            price=5,
            stock=100,
            manufacturer='vishay',
            part_number='MF-1K',
            lifecycle_status='active',
            package_type='SMD-0805',
            parameters={'resistance': '1 кОм'},
        )
        Product.objects.create(
            name='BC547B',
            slug='bc547b',
            category=self.cat_t,
            description='NPN',
            price=9,
            stock=20,
            manufacturer='onsemi',
            part_number='BC547B',
            lifecycle_status='active',
            package_type='TO-92',
            parameters={'type': 'NPN'},
        )

    def test_filters_by_component_type_and_query(self):
        resp = self.client.get(reverse('shop:api_component_search') + '?type=resistor&q=MF')
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['category_slug'], 'resistors')
        self.assertEqual(data['results'][0]['part_number'], 'MF-1K')
        self.assertEqual(data['results'][0]['parameters']['resistance'], '1 кОм')

    def test_npn_uses_transistors_category(self):
        resp = self.client.get(reverse('shop:api_component_search') + '?type=npn&q=BC')
        data = resp.json()
        self.assertEqual(data['category_slug'], 'transistors')
        self.assertEqual(data['results'][0]['part_number'], 'BC547B')


class BomMatchTests(TestCase):
    """api_bom_match — схема → реальные товары из каталога."""

    def setUp(self):
        self.cat_r = Category.objects.create(name='Резисторы', slug='resistors')
        self.cat_c = Category.objects.create(name='Конденсаторы', slug='capacitors')
        self.cheap = Product.objects.create(
            name='Резистор 1к',
            slug='res-1k',
            category=self.cat_r,
            description='1k',
            price=5,
            stock=100,
            manufacturer='vishay',
            part_number='MF-1K',
            lifecycle_status='active',
            parameters={'value': '1k'},
        )
        self.expensive = Product.objects.create(
            name='Резистор 10к premium',
            slug='res-10k',
            category=self.cat_r,
            description='precision',
            price=20,
            stock=100,
            manufacturer='vishay',
            part_number='RC-10K',
            lifecycle_status='active',
        )
        # Out-of-stock — не должен попасть в матч.
        Product.objects.create(
            name='Резистор out-of-stock',
            slug='res-oos',
            category=self.cat_r,
            description='',
            price=1,
            stock=0,
            manufacturer='vishay',
            part_number='OOS',
            lifecycle_status='active',
        )
        # obsolete — тоже отфильтрован (только active/nrnd берём).
        Product.objects.create(
            name='Резистор obsolete',
            slug='res-obs',
            category=self.cat_r,
            description='',
            price=1,
            stock=10,
            manufacturer='vishay',
            part_number='OBS',
            lifecycle_status='obsolete',
        )
        self.cap = Product.objects.create(
            name='Конденсатор 10мкФ',
            slug='cap-10u',
            category=self.cat_c,
            description='',
            price=12,
            stock=50,
            manufacturer='vishay',
            part_number='C-10U',
            lifecycle_status='active',
            parameters={'value': '10u'},
        )

    def _post_match(self, components):
        return self.client.post(
            reverse('shop:api_bom_match'),
            data=json.dumps({'components': components}),
            content_type='application/json',
        )

    def _post_export_xlsx(self, components):
        return self.client.post(
            reverse('shop:api_bom_export_xlsx'),
            data=json.dumps({'components': components}),
            content_type='application/json',
        )

    def test_invalid_json_returns_400(self):
        resp = self.client.post(
            reverse('shop:api_bom_match'),
            data='not json',
            content_type='application/json',
        )
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.json()['ok'])

    def test_groups_by_type_and_counts(self):
        resp = self._post_match(
            [
                {'id': 0, 'type': 'resistor'},
                {'id': 1, 'type': 'resistor'},
                {'id': 2, 'type': 'resistor'},
                {'id': 3, 'type': 'capacitor'},
            ]
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['total_components'], 4)
        by_type = {m['type']: m for m in data['matches']}
        self.assertEqual(by_type['resistor']['count'], 3)
        self.assertEqual(by_type['capacitor']['count'], 1)

    def test_picks_cheapest_active_in_stock(self):
        resp = self._post_match([{'id': 0, 'type': 'resistor'}])
        m = resp.json()['matches'][0]
        # Должен выбрать самый дешёвый из active/in-stock — это «Резистор 1к» по 5 руб.
        self.assertEqual(m['product']['part_number'], 'MF-1K')
        self.assertNotIn(m['product']['part_number'], ('OOS', 'OBS'))

    def test_line_total_and_grand_total(self):
        resp = self._post_match(
            [
                {'id': 0, 'type': 'resistor'},
                {'id': 1, 'type': 'resistor'},
                {'id': 2, 'type': 'capacitor'},
            ]
        )
        data = resp.json()
        # 2 × 5 (резистор) + 1 × 12 (конд) = 22
        self.assertAlmostEqual(data['grand_total'], 22.0, places=2)

    def test_xlsx_export_returns_workbook_with_totals(self):
        resp = self._post_export_xlsx(
            [
                {'id': 0, 'type': 'resistor'},
                {'id': 1, 'type': 'resistor'},
                {'id': 2, 'type': 'capacitor'},
                {'id': 3, 'type': 'unknown_part'},
            ]
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertTrue(resp.content.startswith(b'PK'))

        wb = load_workbook(BytesIO(resp.content), data_only=True)
        ws = wb['BOM']
        values = [
            tuple(cell for cell in row)
            for row in ws.iter_rows(values_only=True)
            if any(cell is not None for cell in row)
        ]
        self.assertTrue(any(row[0] == 'DOLG Bill of Materials' for row in values))
        self.assertTrue(any('MF-1K' in row for row in values))
        self.assertTrue(any('C-10U' in row for row in values))
        self.assertTrue(any('нет в каталоге' in row for row in values))
        total_row = next(row for row in values if 'ИТОГО' in row)
        self.assertEqual(total_row[8], 22)

    def test_alternatives_returned(self):
        resp = self._post_match([{'id': 0, 'type': 'resistor'}])
        m = resp.json()['matches'][0]
        # Резисторов в каталоге два active in-stock: должны попасть в alternatives.
        slugs = {alt['part_number'] for alt in m['alternatives']}
        self.assertIn('MF-1K', slugs)
        self.assertIn('RC-10K', slugs)

    def test_explicit_catalog_ref_is_used_before_type_matching(self):
        resp = self._post_match(
            [
                {'id': 0, 'type': 'resistor', 'catalog_ref': 'RC-10K'},
                {'id': 1, 'type': 'resistor', 'catalog_ref': 'RC-10K'},
            ]
        )
        data = resp.json()
        self.assertEqual(len(data['matches']), 1)
        match = data['matches'][0]
        self.assertEqual(match['product']['part_number'], 'RC-10K')
        self.assertEqual(match['count'], 2)
        self.assertAlmostEqual(match['line_total'], 40.0, places=2)

    def test_bom_warns_about_nominal_mismatch(self):
        resp = self._post_match(
            [
                {'id': 0, 'type': 'resistor', 'resistance': 2000, 'catalog_ref': 'MF-1K'},
            ]
        )
        match = resp.json()['matches'][0]
        self.assertTrue(any('номинал' in warning for warning in match['warnings']))

    def test_bom_warns_about_missing_spice_model(self):
        cat_d = Category.objects.create(name='Диоды', slug='diodes')
        Product.objects.create(
            name='Диод 1N4148',
            slug='1n4148',
            category=cat_d,
            description='',
            price=3,
            stock=100,
            manufacturer='onsemi',
            part_number='1N4148',
            lifecycle_status='active',
        )
        resp = self._post_match(
            [
                {'id': 0, 'type': 'diode', 'catalog_ref': '1N4148'},
            ]
        )
        match = resp.json()['matches'][0]
        self.assertTrue(any('SPICE' in warning for warning in match['warnings']))

    def test_npn_component_maps_to_transistor_category(self):
        cat_t = Category.objects.create(name='Транзисторы', slug='transistors')
        Product.objects.create(
            name='BC547B',
            slug='bc547b',
            category=cat_t,
            description='NPN',
            price=9,
            stock=20,
            manufacturer='onsemi',
            part_number='BC547B',
            lifecycle_status='active',
        )
        resp = self._post_match([{'id': 0, 'type': 'npn'}])
        match = resp.json()['matches'][0]
        self.assertEqual(match['category_slug'], 'transistors')
        self.assertEqual(match['product']['part_number'], 'BC547B')

    def test_unknown_type_returns_no_product(self):
        resp = self._post_match([{'id': 0, 'type': 'mystery_widget'}])
        m = resp.json()['matches'][0]
        self.assertIsNone(m['product'])
        self.assertEqual(m['line_total'], 0.0)

    def test_components_must_be_list(self):
        resp = self.client.post(
            reverse('shop:api_bom_match'),
            data=json.dumps({'components': 'not a list'}),
            content_type='application/json',
        )
        self.assertEqual(resp.status_code, 400)


class BomAddAllTests(TestCase):
    """api_bom_add_all — массовое добавление в корзину."""

    def setUp(self):
        cat = Category.objects.create(name='Резисторы', slug='resistors')
        self.p = Product.objects.create(
            name='Резистор 1к',
            slug='r-1k',
            category=cat,
            description='',
            price=5,
            stock=100,
            manufacturer='vishay',
            part_number='MF-1K',
            lifecycle_status='active',
        )

    def test_creates_cart_items(self):
        resp = self.client.post(
            reverse('shop:api_bom_add_all'),
            data=json.dumps({'items': [{'slug': 'r-1k', 'quantity': 3}]}),
            content_type='application/json',
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()['ok'])
        self.assertEqual(resp.json()['added'], 1)
        self.assertIn('project_cart', resp.json())

    def test_increments_existing_item(self):
        # Первый раз — создаст. Второй — увеличит количество.
        for _ in range(2):
            self.client.post(
                reverse('shop:api_bom_add_all'),
                data=json.dumps({'items': [{'slug': 'r-1k', 'quantity': 2}]}),
                content_type='application/json',
            )
        from shop.models import CartItem

        items = CartItem.objects.filter(product=self.p)
        self.assertEqual(items.count(), 1)
        self.assertEqual(items.first().quantity, 4)

    def test_skips_unknown_slug(self):
        resp = self.client.post(
            reverse('shop:api_bom_add_all'),
            data=json.dumps(
                {
                    'items': [
                        {'slug': 'nonexistent', 'quantity': 1},
                        {'slug': 'r-1k', 'quantity': 1},
                    ]
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(resp.json()['added'], 1)

    def test_empty_items_returns_400(self):
        resp = self.client.post(
            reverse('shop:api_bom_add_all'),
            data=json.dumps({'items': []}),
            content_type='application/json',
        )
        self.assertEqual(resp.status_code, 400)

    def test_get_method_rejected(self):
        resp = self.client.get(reverse('shop:api_bom_add_all'))
        self.assertEqual(resp.status_code, 405)

    def test_caps_quantity_by_stock(self):
        resp = self.client.post(
            reverse('shop:api_bom_add_all'),
            data=json.dumps({'items': [{'slug': 'r-1k', 'quantity': 150}]}),
            content_type='application/json',
        )
        from shop.models import CartItem

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(CartItem.objects.get(product=self.p).quantity, 100)
        self.assertEqual(resp.json()['added'], 1)
        self.assertEqual(len(resp.json()['limited']), 1)

    def test_project_cart_context_is_saved_and_rendered(self):
        from Dolg_APP.models import SchematicProject

        user = get_user_model().objects.create_user(username='engineer', password='pass')
        self.client.force_login(user)
        project = SchematicProject.objects.create(
            user=user,
            name='RC filter BOM demo',
            description='BOM source project',
            scheme_data={'components': []},
        )
        resp = self.client.post(
            reverse('shop:api_bom_add_all'),
            data=json.dumps(
                {
                    'items': [{'slug': 'r-1k', 'quantity': 2}],
                    'project': {'id': project.id, 'name': 'wrong client name'},
                    'source': 'simulation',
                }
            ),
            content_type='application/json',
        )

        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data['project_cart']['project_name'], project.name)
        self.assertEqual(data['project_cart']['total_quantity'], 2)

        cart_response = self.client.get(reverse('shop:cart'))
        self.assertContains(cart_response, 'Проектная корзина')
        self.assertContains(cart_response, project.name)
        self.assertContains(cart_response, self.p.part_number)

    def test_project_cart_context_can_be_cleared(self):
        session = self.client.session
        session['project_cart_context'] = {'project_name': 'Temporary project'}
        session.save()

        response = self.client.post(reverse('shop:clear_project_cart_context'))

        self.assertRedirects(response, reverse('shop:cart'))
        self.assertNotIn('project_cart_context', self.client.session)


class CompareToggleTests(TestCase):
    """compare_toggle — добавление/удаление товара в сессионное сравнение."""

    def setUp(self):
        cat = Category.objects.create(name='ICs', slug='ics')
        self.p1 = Product.objects.create(
            name='STM32',
            slug='stm32',
            category=cat,
            description='',
            price=120,
            stock=10,
            manufacturer='st',
            part_number='STM32',
        )
        self.p2 = Product.objects.create(
            name='ATMega',
            slug='atmega',
            category=cat,
            description='',
            price=80,
            stock=10,
            manufacturer='atmel',
            part_number='ATMEGA',
        )

    def test_toggle_adds_then_removes(self):
        self.client.post(reverse('shop:compare_toggle', args=['stm32']))
        self.assertIn('stm32', self.client.session.get('compare', []))
        # Повторный POST — снимает.
        self.client.post(reverse('shop:compare_toggle', args=['stm32']))
        self.assertNotIn('stm32', self.client.session.get('compare', []))

    def test_compare_clear(self):
        self.client.post(reverse('shop:compare_toggle', args=['stm32']))
        self.client.post(reverse('shop:compare_toggle', args=['atmega']))
        self.client.post(reverse('shop:compare_clear'))
        self.assertEqual(self.client.session.get('compare', []), [])
